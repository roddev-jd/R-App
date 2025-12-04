# Importaciones estándar
import asyncio
import io
import json
import logging
import os
import tempfile
from collections import deque
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse, quote
from typing import List, Dict, Any, Optional, Callable, Iterator
from concurrent.futures import ThreadPoolExecutor

# Importaciones de terceros
# DuckDB es opcional - sistema funcionará sin él usando fallback a Pandas
try:
    import duckdb
    DUCKDB_AVAILABLE = True
    logging.info("DuckDB disponible - modo completo habilitado")
except ImportError:
    DUCKDB_AVAILABLE = False
    duckdb = None
    logging.warning("DuckDB no disponible - usando modo legacy con Pandas puro")

import pandas as pd
import requests
import xlsxwriter
from azure.storage.blob import BlobServiceClient

# Imports de servicios de utilidades
from services.csv_utils import csv_utils
from services.dataframe_utils import dataframe_utils
from services.storage_utils import (_download_file_from_download_url,
                                  _parse_sharepoint_url, _download_blob_with_progress)
from services.storage_utils import _extract_container_name_from_url as extract_container_name_pure
from services.filter_service import (_get_empty_filter_response, _get_sku_column_candidates,
                                   _extract_filter_options)
from services.data_service import (_create_config_parser, _parse_filter_section,
                                 _build_blob_config, get_blob_config, get_dynamic_config_path)
from services.progress_utils import DataLoadProgressTracker
# Cache system simplified - using only persistent cache

# Importaciones locales
from services import sharepoint_service as sharepoint_auth
from core.sse_channel import search_progress_queue, clear_data_load_progress_queue
from core.utils import getenv_int
from services.cache_service import persistent_cache

# FASE 2.1: Async storage utilities (NEW)
try:
    from services.async_storage_utils import (
        download_from_azure_blob_async,
        download_blob_with_progress_async,
        check_async_availability
    )
    from services.async_sharepoint_service import (
        download_sharepoint_with_progress_async,
        read_excel_from_sharepoint_async,
        read_csv_from_sharepoint_async
    )
    ASYNC_STORAGE_AVAILABLE = True
    logging.info("✅ Módulos async de storage disponibles")
except ImportError as e:
    ASYNC_STORAGE_AVAILABLE = False
    logging.warning(f"⚠️ Módulos async de storage no disponibles: {e}")
    logging.warning("Sistema funcionará en modo síncrono legacy")

# Configuración de logging
logging.basicConfig(
    level=logging.INFO, 
    format='[%(asctime)s] [%(levelname)s] (%(module)s:%(lineno)d) %(message)s', 
    datefmt='%H:%M:%S'
)

# Constantes y configuración global
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR = Path.home() / ".cache" / "ReportesRodrobus"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# Thread pools especializados
io_executor = ThreadPoolExecutor(max_workers=8, thread_name_prefix="io_worker")
sharepoint_executor = ThreadPoolExecutor(max_workers=4, thread_name_prefix="sp_worker")

# Cachés globales eliminados - caché persistente maneja todo


# Configuración de archivos
CONFIG_FILE_PATH = get_dynamic_config_path(SCRIPT_DIR)
CONFIG_FILE_NAME = os.path.basename(CONFIG_FILE_PATH)

# Funciones de descarga y conectores


# Funciones de limpieza de datos



# Funciones de conectores externos





# Sistema de caché y logging
MAX_LOG_QUEUE_SIZE = getenv_int("MAX_LOG_QUEUE_SIZE", 500)
LOG_CLEANUP_THRESHOLD = int(MAX_LOG_QUEUE_SIZE * 0.8)
# TTL Cache configuration removed - using persistent cache only

# Cola de logs para SSE
log_queue = deque(maxlen=MAX_LOG_QUEUE_SIZE)
new_log_event = asyncio.Event()

# Cache system eliminated - using only persistent cache for all databases

# Cache control eliminated - persistent cache handles all caching

def cleanup_old_logs():
    """Limpia logs antiguos manteniendo solo los relevantes."""
    if len(log_queue) < LOG_CLEANUP_THRESHOLD:
        return
        
    important_types = {"download_progress", "export_progress", "error"}
    recent_logs = []
    
    for log_entry in reversed(log_queue):
        try:
            log_data = json.loads(log_entry)
            if (log_data.get("type") in important_types or 
                len(recent_logs) < MAX_LOG_QUEUE_SIZE // 2):
                recent_logs.insert(0, log_entry)
                
            if len(recent_logs) >= MAX_LOG_QUEUE_SIZE // 2:
                break
        except json.JSONDecodeError:
            recent_logs.insert(0, log_entry)
    
    log_queue.clear()
    log_queue.extend(recent_logs)
    logging.debug(f"Log queue limpiada: {len(recent_logs)} logs mantenidos")

# Cache expiration function eliminated - persistent cache handles all caching




class QueueLoggingHandler(logging.Handler):
    def __init__(self, queue_instance):
        super().__init__()
        self.queue_instance = queue_instance
        self.setFormatter(logging.Formatter('[%(asctime)s] [%(levelname)s] (%(module)s) %(message)s', datefmt='%H:%M:%S'))

    def emit(self, record):
        log_entry = self.format(record)
        self.queue_instance.append(log_entry)
        try:
            new_log_event.set()
        except RuntimeError: # Puede ocurrir si el loop de eventos no está corriendo
            pass
        new_log_event.clear()

def setup_log_streaming():
    root_logger = logging.getLogger()
    if root_logger.level > logging.INFO:
        root_logger.setLevel(logging.INFO)
    queue_handler = QueueLoggingHandler(log_queue)
    if not any(isinstance(h, QueueLoggingHandler) for h in root_logger.handlers):
        root_logger.addHandler(queue_handler)
        logging.info("Manejador de logging para streaming SSE configurado.")
    else:
        logging.info("Manejador de logging para streaming SSE ya estaba configurado.")

# ... (Variables de config_data sin cambios) ...
config_data: Dict[str, Any] = {
    "config_parser": None,
    "connection_string": None,
    "container_url": None,
    "container_name": None,
    "default_blob_filename": None,
    "blob_options": {},
    "filter_configs": {},
    "saved_states": {},
    "local_file_base_path": SCRIPT_DIR
}

current_blob_display_name: Optional[str] = None
df_original: pd.DataFrame = pd.DataFrame()
df_filtered: pd.DataFrame = pd.DataFrame()
# Tipo Any usado para compatibilidad cuando DuckDB no está disponible
duckdb_conn: Optional[Any] = None
sku_hijo_filter_list: Optional[List[str]] = None
sku_padre_filter_list: Optional[List[str]] = None
ticket_filter_list: Optional[List[str]] = None # NUEVA VARIABLE GLOBAL

# Variable global para el buscador integral de carpetas
folder_search_excel_data: Optional[bytes] = None

# Configuración del sitio SharePoint por defecto
DEFAULT_SHAREPOINT_SITE = "https://ripleycorp.sharepoint.com/sites/Publicacinycontenido2"

# Configuración del sistema

def _load_azure_config(parser):
    """Carga configuración de Azure desde variables de entorno o archivo.

    Prioridad de lectura:
    1. Variables de entorno (AZURE_CONNECTION_STRING, AZURE_CONTAINER_URL, AZURE_DEFAULT_BLOB)
    2. Sección [Azure] en config.ini (fallback)
    """
    has_azure_section = parser.has_section('Azure')

    # 1. Obtener ConnectionString - priorizar variable de entorno
    temp_connection_string = os.environ.get("AZURE_CONNECTION_STRING")
    if temp_connection_string:
        logging.info(f"ConnectionString obtenida desde variable de entorno AZURE_CONNECTION_STRING (len={len(temp_connection_string)}).")
    elif has_azure_section:
        # Fallback a config.ini
        conn_from_file = parser.get('Azure', 'ConnectionString', fallback='').strip()
        if conn_from_file:
            logging.info(f"Usando ConnectionString desde config.ini (len={len(conn_from_file)}).")
            temp_connection_string = conn_from_file

    if not temp_connection_string:
        logging.warning("No se encontró AZURE_CONNECTION_STRING en variables de entorno ni en config.ini.")
        return

    # 2. Obtener ContainerURL - priorizar variable de entorno
    temp_container_url = os.environ.get("AZURE_CONTAINER_URL")
    if temp_container_url:
        logging.info(f"ContainerURL obtenida desde variable de entorno AZURE_CONTAINER_URL.")
    elif has_azure_section:
        temp_container_url = parser.get('Azure', 'ContainerURL', fallback=None)
        if temp_container_url:
            logging.info(f"Usando ContainerURL desde config.ini.")

    if not temp_container_url:
        logging.warning("No se encontró AZURE_CONTAINER_URL en variables de entorno ni en config.ini.")
        return

    # 3. Obtener DefaultBlob - priorizar variable de entorno (opcional)
    default_blob = os.environ.get("AZURE_DEFAULT_BLOB")
    if default_blob:
        logging.info(f"DefaultBlob obtenido desde variable de entorno AZURE_DEFAULT_BLOB: {default_blob}")
    elif has_azure_section:
        default_blob = parser.get('Azure', 'DefaultBlob', fallback=None)
        if default_blob:
            logging.info(f"Usando DefaultBlob desde config.ini: {default_blob}")

    config_data["default_blob_filename"] = default_blob

    # 4. Configurar Azure con los valores obtenidos
    config_data["connection_string"] = temp_connection_string
    config_data["container_url"] = temp_container_url

    # Extraer nombre del contenedor
    parsed_url = urlparse(config_data["container_url"])
    path_parts = parsed_url.path.strip('/').split('/')
    if path_parts and path_parts[0]:
        config_data["container_name"] = path_parts[0]
        logging.info(f"Contenedor Azure configurado: '{config_data['container_name']}'")
    else:
        _extract_container_name_from_url()

def _extract_container_name_from_url():
    """Extrae el nombre del contenedor de URL complejas."""
    name_found = extract_container_name_pure(config_data["container_url"])
    config_data["container_name"] = name_found

def _load_blob_options(parser):
    """Carga opciones de blobs desde configuración."""
    if not parser.has_section('Blobs'):
        config_data["blob_options"] = {}
        return
        
    temp_blob_options = {}
    suffixes = ['_display_name', '_source_type', '_country', '_description',
                '_filter_config_ref', '_local_path', '_use_local', '_prefilter_nom_estado',
                '_enrichment_source', '_enrichment_join_column', '_enrichment_columns',
                '_sheet_name', '_values_only', '_file_pattern']
    
    # Encontrar claves base
    all_keys = parser.options('Blobs')
    base_keys = [k for k in all_keys if not any(k.endswith(suffix) for suffix in suffixes)]
    
    for key in base_keys:
        display_name = parser.get('Blobs', f"{key}_display_name", fallback=key)
        temp_blob_options[display_name.upper()] = _build_blob_config(parser, key, display_name)
    
    config_data["blob_options"] = temp_blob_options
    logging.info(f"Cargadas {len(temp_blob_options)} opciones de blobs.")

# get_blob_config() migrada a data_service.py - ahora usa config_data como parámetro
def get_blob_config_wrapper(display_name: str) -> Optional[Dict[str, Any]]:
    """Wrapper para get_blob_config() que pasa config_data."""
    return get_blob_config(display_name, config_data)



def _load_filter_configs(parser):
    """Carga configuraciones de filtros para cada blob."""
    for blob_key_upper, blob_attrs in config_data["blob_options"].items():
        filter_section_key = blob_attrs["filter_config_ref"]
        filter_section_name = f"Filters_{filter_section_key}"
        
        f_config = {
            'filter_cols': [], 'hide_cols': [], 'not_empty_cols': [], 
            'hide_values': {}, 'exclude_rows': {}
        }
        
        if parser.has_section(filter_section_name):
            f_config.update(_parse_filter_section(parser, filter_section_name))
        
        config_data["filter_configs"][blob_key_upper] = f_config


def _load_saved_states(parser):
    """Carga estados guardados para cada blob."""
    for display_name_key_upper in config_data["blob_options"].keys():
        saved_state_section_name = f"SavedState_{display_name_key_upper}"
        s_state = {}
        
        if parser.has_section(saved_state_section_name):
            for key_opt in parser.options(saved_state_section_name):
                s_state[key_opt] = parser.get(saved_state_section_name, key_opt, fallback='')
        
        config_data["saved_states"][display_name_key_upper] = s_state

def load_app_config() -> bool:
    """Carga la configuración principal de la aplicación."""
    global config_data
    try:
        parser = _create_config_parser()
        
        if not os.path.exists(CONFIG_FILE_PATH):
            logging.error(f"No se encontró config.ini en: {CONFIG_FILE_PATH}")
            return False
        
        parser.read(CONFIG_FILE_PATH, encoding='utf-8')
        config_data["config_parser"] = parser

        _load_azure_config(parser)
        _load_blob_options(parser)
        _load_filter_configs(parser)
        _load_saved_states(parser)

        setup_log_streaming()
        logging.info("Configuración de la aplicación cargada y log streaming configurado.")
        return True

    except Exception as e:
        logging.error(f"Error fatal al cargar configuración de la aplicación: {e}", exc_info=True)
        return False





# Funciones de carga de datos


def _setup_duckdb_connection(df: pd.DataFrame):
    """Configura nueva conexión DuckDB con los datos."""
    global duckdb_conn

    # Si DuckDB no está disponible, salir temprano
    if not DUCKDB_AVAILABLE:
        logging.info("DuckDB no disponible - saltando configuración (modo legacy)")
        return

    logging.info(f"Iniciando configuración DuckDB - Estado actual: {duckdb_conn is not None}")

    # Validar DataFrame de entrada
    if df is None:
        logging.error("Error: DataFrame es None")
        raise ValueError("DataFrame no puede ser None")

    if df.empty:
        logging.warning("DataFrame está vacío, pero continuando...")

    try:
        # Cerrar conexión anterior si existe
        if duckdb_conn is not None:
            try:
                duckdb_conn.close()
                logging.info("Conexión DuckDB anterior cerrada")
            except Exception as e:
                logging.warning(f"Error cerrando conexión anterior: {e}")
            finally:
                duckdb_conn = None

        # Crear nueva conexión
        duckdb_conn = duckdb.connect(database=':memory:')
        logging.info("Nueva conexión DuckDB creada")

        # Registrar DataFrame
        duckdb_conn.register('pandas_df', df)
        logging.info(f"DataFrame registrado con {len(df)} filas")

        # Crear tabla
        duckdb_conn.execute("CREATE OR REPLACE TABLE data AS SELECT * FROM pandas_df")
        logging.info("Tabla 'data' creada exitosamente")

        # Limpiar registro temporal
        duckdb_conn.unregister('pandas_df')

        # Validación crítica: verificar que la conexión funciona
        if duckdb_conn is None:
            raise RuntimeError("Error crítico: duckdb_conn es None después de la configuración")

        # Test de conectividad
        result = duckdb_conn.execute("SELECT COUNT(*) FROM data").fetchone()
        expected_count = len(df)
        actual_count = result[0] if result else 0

        if actual_count != expected_count:
            raise RuntimeError(f"Error de consistencia: esperado {expected_count}, obtenido {actual_count}")

        logging.info(f"✅ Configuración DuckDB exitosa - {actual_count} filas disponibles")

    except Exception as e:
        logging.error(f"❌ Error en configuración DuckDB: {e}")
        # Cleanup en caso de error
        if duckdb_conn is not None:
            try:
                duckdb_conn.close()
            except:
                pass
            duckdb_conn = None
        raise

def _load_from_persistent_cache(param_from_frontend_url: str, selected_columns: Optional[List[str]] = None) -> Optional[Dict[str, Any]]:
    """
    Intenta cargar datos desde caché persistente.

    Args:
        param_from_frontend_url: Nombre de la base de datos
        selected_columns: Lista opcional de columnas a cargar (optimización de RAM con Parquet)
    """
    if not persistent_cache.has_cached_data(param_from_frontend_url):
        return None

    try:
        # Log inicial (sin enviar SSE todavía)
        if selected_columns:
            logging.info(f"Cargando '{param_from_frontend_url}' desde caché persistente ({len(selected_columns)} columnas seleccionadas).")
        else:
            logging.info(f"Cargando '{param_from_frontend_url}' desde caché persistente.")

        # Limpiar estado anterior
        global df_original, df_filtered, current_blob_display_name, duckdb_conn
        df_original = pd.DataFrame()
        df_filtered = pd.DataFrame()
        current_blob_display_name = None

        if duckdb_conn:
            try:
                duckdb_conn.close()
                logging.info("Conexión DuckDB anterior cerrada.")
            except Exception as e:
                logging.warning(f"Error cerrando DuckDB: {e}")
            duckdb_conn = None

        # Intentar cargar datos del caché (sin tracker aún)
        df_original = persistent_cache.load_cached_data(param_from_frontend_url, columns=selected_columns)

        # Si la carga falló, retornar sin enviar mensajes SSE
        if df_original is None:
            return None

        # SOLO AHORA crear el tracker de progreso (datos confirmados)
        progress_tracker = DataLoadProgressTracker(
            operation_name="Carga desde caché",
            blob_display_name=param_from_frontend_url
        )

        # Enviar mensajes de progreso ahora que sabemos que hay datos
        if selected_columns:
            progress_tracker.update_progress(20, "cache", f"Cargando {len(selected_columns)} columnas desde caché local...")
        else:
            progress_tracker.update_progress(20, "cache", "Cargando datos desde caché local...")

        progress_tracker.update_progress(60, "cache", f"Datos cargados: {len(df_original):,} registros")

        current_blob_display_name = param_from_frontend_url
        _setup_duckdb_connection(df_original)

        # Obtener configuración para filtros
        progress_tracker.update_progress(80, "processing", "Preparando filtros...")
        blob_config = get_blob_config_wrapper(param_from_frontend_url)
        if not blob_config:
            return None

        filter_options = _extract_filter_options(df_original, blob_config, config_data["filter_configs"])

        logging.info(f"Estado sincronizado para '{param_from_frontend_url}'. Shape: {df_original.shape}")

        # Enviar mensaje de completado
        progress_tracker.finish(success=True, final_message=f"Carga completada desde caché: {len(df_original):,} registros")

        return {
            "message": f"Datos de '{param_from_frontend_url}' cargados desde caché persistente.",
            "row_count_original": len(df_original),
            "columns": list(df_original.columns),
            "filter_options": filter_options,
            "source_type": blob_config.get('source_type', 'desconocido'),
            "from_cache": True
        }
        
    except Exception as e:
        logging.error(f"Error cargando desde caché persistente '{param_from_frontend_url}': {e}", exc_info=True)
        # Limpiar caché corrupto
        persistent_cache.clear_cache(param_from_frontend_url)
        return None

def load_blob_data(param_from_frontend_url: str, selected_columns_from_api: Optional[List[str]] = None) -> Dict[str, Any]:
    """Carga datos de blob con verificación inteligente de caché.

    CACHÉ INTELIGENTE:
    - Para fuentes SharePoint: verifica automáticamente si hay actualizaciones antes de usar caché
    - Si hay actualización: descarga la versión nueva
    - Si NO hay actualización: usa caché local (rápido)
    - Para otras fuentes: usa caché sin verificación

    Args:
        param_from_frontend_url: Identificador de la fuente de datos
        selected_columns_from_api: Lista opcional de columnas específicas desde API (prioridad sobre config)

    Returns:
        Diccionario con información de la carga realizada, incluyendo:
        - cache_decision: 'using_cache' | 'downloading_fresh' | 'no_cache'
    """
    import time
    start_time = time.time()
    global df_original, current_blob_display_name, config_data, duckdb_conn

    # Limpiar mensajes antiguos del queue de SSE antes de iniciar nueva carga
    clear_data_load_progress_queue()

    # Crear tracker de progreso para reportar al frontend via SSE
    progress_tracker = DataLoadProgressTracker(
        operation_name="Carga de datos",
        blob_display_name=param_from_frontend_url
    )
    progress_tracker.update_progress(1, "init", f"Iniciando carga de '{param_from_frontend_url}'...")

    logging.info(f"load_blob_data: Solicitud para '{param_from_frontend_url}'")

    # Limpiar filtros globales
    global sku_hijo_filter_list, sku_padre_filter_list, ticket_filter_list
    sku_hijo_filter_list = None
    sku_padre_filter_list = None
    ticket_filter_list = None
    logging.info(f"Filtros globales limpiados para '{param_from_frontend_url}'.")

    # Determinar columnas a cargar: API tiene prioridad sobre config
    selected_columns_check = None
    if selected_columns_from_api:
        selected_columns_check = [col.strip().lower() for col in selected_columns_from_api]
        logging.info(f"Usando columnas desde API: {len(selected_columns_check)} columnas")
    else:
        # Leer configuración de selectcolumns del config.ini
        key_for_config_check = param_from_frontend_url.upper()
        config_settings_check = config_data["filter_configs"].get(key_for_config_check, {})
        select_columns_check = config_settings_check.get('selectcolumns', '')  # Note: sin guión bajo en config.ini
        if select_columns_check and select_columns_check.strip():
            selected_columns_check = [col.strip().lower() for col in select_columns_check.split(',') if col.strip()]
            logging.info(f"Usando columnas desde config: {len(selected_columns_check)} columnas")

    # CACHÉ INTELIGENTE: Verificar actualizaciones automáticamente antes de usar caché
    cache_decision = "no_cache"  # Valores: "using_cache", "downloading_fresh", "no_cache"

    # ✅ HITO 1.3: Sistema de información de verificación de caché
    cache_verification_status = {
        'verified': False,
        'status': 'not_cached',  # 'verified_fresh' | 'verified_stale' | 'verification_failed' | 'expired_by_age' | 'not_cached'
        'message': '',
        'last_check': None,
        'error_detail': None
    }

    if persistent_cache.is_cacheable(param_from_frontend_url) and persistent_cache.has_cached_data(param_from_frontend_url):
        # Obtener configuración de la base de datos
        key_for_blob_lookup = param_from_frontend_url.upper()
        found_blob_attrs = config_data.get("blob_options", {}).get(key_for_blob_lookup)

        if found_blob_attrs:
            source_type = found_blob_attrs.get("source_type")

            # ✅ HITO 2.3: Verificar expiración por edad ANTES de verificación remota
            if persistent_cache.is_cache_expired(param_from_frontend_url):
                logging.info(f"⏰ Caché expirado por política de edad - Descargando fresco")
                persistent_cache.clear_cache(param_from_frontend_url)
                cache_decision = "downloading_fresh"
                cache_verification_status = {
                    'verified': False,
                    'status': 'expired_by_age',
                    'message': f'Caché excedió edad máxima de {persistent_cache.CACHE_MAX_AGE_DAYS} días',
                    'last_check': datetime.now(timezone.utc).isoformat()
                }
            # Verificar actualizaciones para fuentes SharePoint y Azure (HITO 1.4)
            elif source_type in ["sharepoint", "azure"]:
                try:
                    progress_tracker.update_progress(2, "verifying", "Verificando si hay actualizaciones disponibles...")

                    if source_type == "sharepoint":
                        # SharePoint: usar Graph API (código existente)
                        auth = get_sharepoint_authenticator()
                        access_token = auth.get_token()
                        auth_headers = {"Authorization": f"Bearer {access_token}"}
                        source_url = found_blob_attrs.get("source_url", "")

                        update_info = persistent_cache.check_remote_update(
                            param_from_frontend_url,
                            source_url,
                            auth_headers,
                            source_type="sharepoint"
                        )

                    elif source_type == "azure":
                        # Azure: usar blob properties (NUEVO - HITO 1.4)
                        logging.info(f"Preparando verificación Azure para '{param_from_frontend_url}'...")

                        # Obtener configuración de Azure desde config_data
                        filename = found_blob_attrs.get("filename")
                        conn_str = config_data.get("connection_string")
                        azure_config = {
                            'connection_string': conn_str,
                            'container_name': config_data.get("container_name"),
                            'blob_name': filename
                        }

                        # Debug: verificar si connection_string está presente
                        if conn_str:
                            logging.info(f"Azure config: container='{azure_config['container_name']}', blob='{azure_config['blob_name']}', conn_str_len={len(conn_str)}")
                        else:
                            logging.error(f"Azure config: connection_string es None o vacía. container='{azure_config['container_name']}', blob='{azure_config['blob_name']}'")

                        update_info = persistent_cache.check_remote_update(
                            param_from_frontend_url,
                            source_url="",  # No necesario para Azure
                            auth_headers=None,  # Azure usa connection string
                            source_type="azure",
                            azure_config=azure_config
                        )

                    # ✅ HITO 1.3: LÓGICA COMÚN con sistema de advertencias
                    if update_info.get('error'):
                        # Si hay error verificando, usar caché como fallback seguro
                        cache_decision = "using_cache"
                        cache_verification_status = {
                            'verified': False,
                            'status': 'verification_failed',
                            'message': f"⚠️ No se pudo verificar actualización: {update_info['error']}. Usando caché local sin verificar.",
                            'last_check': datetime.now(timezone.utc).isoformat(),
                            'error_detail': update_info.get('error'),
                            'comparison_details': update_info.get('comparison_details', '')
                        }
                        logging.warning(
                            f"⚠️ Error verificando actualizaciones para '{param_from_frontend_url}': "
                            f"{update_info['error']} - Usando caché SIN VERIFICAR"
                        )
                    elif update_info.get('update_available'):
                        # Hay actualización disponible: limpiar caché y descargar fresco
                        persistent_cache.clear_cache(param_from_frontend_url)
                        cache_decision = "downloading_fresh"
                        cache_verification_status = {
                            'verified': True,
                            'status': 'verified_stale',
                            'message': f"Actualización disponible - descargando datos frescos",
                            'last_check': datetime.now(timezone.utc).isoformat(),
                            'reason': update_info.get('reason', 'Datos remotos más recientes')
                        }
                        logging.info(
                            f"🔄 Actualización detectada para '{param_from_frontend_url}' - "
                            f"Descargando versión nueva. {update_info.get('comparison_details', '')}"
                        )
                    else:
                        # Caché está actualizado: usar versión local
                        cache_decision = "using_cache"
                        cache_verification_status = {
                            'verified': True,
                            'status': 'verified_fresh',
                            'message': f"✅ Caché verificado como actualizado al {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                            'last_check': datetime.now(timezone.utc).isoformat(),
                            'reason': update_info.get('reason', 'Caché coincide con fuente remota')
                        }
                        logging.info(
                            f"✅ Caché de '{param_from_frontend_url}' VERIFICADO como actualizado - "
                            f"Usando versión local. {update_info.get('comparison_details', '')}"
                        )

                except Exception as e:
                    # Fallback: usar caché en caso de cualquier error
                    cache_decision = "using_cache"
                    cache_verification_status = {
                        'verified': False,
                        'status': 'verification_failed',
                        'message': f"⚠️ Error inesperado durante verificación: {str(e)}. Usando caché local sin verificar.",
                        'last_check': datetime.now(timezone.utc).isoformat(),
                        'error_detail': str(e)
                    }
                    logging.warning(
                        f"⚠️ Error verificando actualizaciones para '{param_from_frontend_url}': {e} - "
                        f"Usando caché SIN VERIFICAR"
                    )

            elif source_type == "local_partitioned_csv":
                # Verificar actualizaciones para CSV particionados locales
                try:
                    progress_tracker.update_progress(2, "verifying", "Verificando si hay actualizaciones en archivos locales...")

                    logging.info(f"Verificando archivos particionados para '{param_from_frontend_url}'...")

                    # Obtener configuración
                    base_directory = found_blob_attrs.get("value")  # value contiene la ruta del directorio
                    file_pattern = found_blob_attrs.get("file_pattern", "*.csv")

                    # Preparar configuración para verificación
                    partitioned_config = {
                        'base_directory': base_directory,
                        'file_pattern': file_pattern
                    }

                    # Llamar a check_remote_update (que ahora soporta local_partitioned_csv)
                    update_info = persistent_cache.check_remote_update(
                        param_from_frontend_url,
                        source_url="",  # No usado para archivos locales
                        auth_headers=None,  # No usado para archivos locales
                        source_type="local_partitioned_csv",
                        azure_config=partitioned_config  # Reutilizamos este parámetro para pasar config
                    )

                    # ✅ LÓGICA COMÚN: Mismo manejo que SharePoint/Azure
                    if update_info.get('error'):
                        # Si hay error verificando, usar caché como fallback seguro
                        cache_decision = "using_cache"
                        cache_verification_status = {
                            'verified': False,
                            'status': 'verification_failed',
                            'message': f"⚠️ No se pudo verificar actualización: {update_info['error']}. Usando caché local sin verificar.",
                            'last_check': datetime.now(timezone.utc).isoformat(),
                            'error_detail': update_info.get('error'),
                            'comparison_details': update_info.get('comparison_details', '')
                        }
                        logging.warning(
                            f"⚠️ Error verificando archivos particionados para '{param_from_frontend_url}': "
                            f"{update_info['error']} - Usando caché SIN VERIFICAR"
                        )
                    elif update_info.get('update_available'):
                        # Hay actualización disponible: limpiar caché y descargar fresco
                        persistent_cache.clear_cache(param_from_frontend_url)
                        cache_decision = "downloading_fresh"
                        cache_verification_status = {
                            'verified': True,
                            'status': 'verified_stale',
                            'message': f"Actualización disponible en archivos locales - recargando datos frescos",
                            'last_check': datetime.now(timezone.utc).isoformat(),
                            'reason': update_info.get('reason', 'Archivos locales modificados')
                        }
                        logging.info(
                            f"🔄 Actualización detectada en archivos particionados '{param_from_frontend_url}' - "
                            f"Recargando versión nueva. {update_info.get('comparison_details', '')}"
                        )
                    else:
                        # Caché está actualizado: usar versión local
                        cache_decision = "using_cache"
                        cache_verification_status = {
                            'verified': True,
                            'status': 'verified_fresh',
                            'message': f"✅ Caché verificado como actualizado al {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                            'last_check': datetime.now(timezone.utc).isoformat(),
                            'reason': update_info.get('reason', 'Caché coincide con archivos locales')
                        }
                        logging.info(
                            f"✅ Caché de archivos particionados '{param_from_frontend_url}' VERIFICADO como actualizado - "
                            f"Usando versión local. {update_info.get('comparison_details', '')}"
                        )

                except Exception as e:
                    # Fallback: usar caché en caso de cualquier error
                    cache_decision = "using_cache"
                    cache_verification_status = {
                        'verified': False,
                        'status': 'verification_failed',
                        'message': f"⚠️ Error inesperado durante verificación: {str(e)}. Usando caché local sin verificar.",
                        'last_check': datetime.now(timezone.utc).isoformat(),
                        'error_detail': str(e)
                    }
                    logging.warning(
                        f"⚠️ Error verificando archivos particionados para '{param_from_frontend_url}': {e} - "
                        f"Usando caché SIN VERIFICAR"
                    )

            else:
                # Fuentes no-SharePoint y no-Azure (FTP, CSV local, etc.): usar caché sin verificación
                cache_decision = "using_cache"
                logging.info(f"Fuente tipo '{source_type}' no soporta verificación de actualizaciones - usando caché si existe")

    # Intentar cargar desde caché persistente (si no se limpió por actualización)
    cached_result = _load_from_persistent_cache(param_from_frontend_url, selected_columns=selected_columns_check)
    if cached_result:
        # ✅ HITO 1.3: Agregar información de verificación al resultado
        cached_result["cache_decision"] = cache_decision
        cached_result["cache_info"] = cache_verification_status
        return cached_result

    # Si no hay cache, proceder con descarga normal desde la fuente original
    if cache_decision == "no_cache":
        logging.info(f"Datos para '{param_from_frontend_url}' no encontrados en caché. Cargando desde fuente...")
    else:
        logging.info(f"Caché limpiado por actualización. Descargando desde fuente...")

    key_for_blob_options_lookup = param_from_frontend_url.upper()
    found_blob_attrs = config_data.get("blob_options", {}).get(key_for_blob_options_lookup)

    if not found_blob_attrs:
        raise ValueError(f"Fuente de datos no encontrada para el display_name: {param_from_frontend_url}")

    filename = found_blob_attrs.get("value")
    source_type = found_blob_attrs.get("source_type")

    # Crear mensaje específico según el tipo de fuente
    source_message_map = {
        "azure_blob": "Descargando desde Azure Blob Storage...",
        "sharepoint": "Descargando desde SharePoint...",
        "ftp": "Descargando desde servidor FTP...",
        "s3": "Descargando desde Amazon S3...",
        "local": "Cargando desde archivo local..."
    }
    download_message = source_message_map.get(source_type, f"Descargando desde {source_type}...")
    progress_tracker.update_progress(5, "download", download_message)

    # --- Soporte opcional de carga local ---
    use_local_flag = found_blob_attrs.get("use_local")
    if use_local_flag:
        local_path_cfg = found_blob_attrs.get("local_path")
        if not local_path_cfg:
            raise FileNotFoundError(
                f"Se indicó 'use_local' para '{param_from_frontend_url}', pero no se configuró la ruta local (_local_path)."
            )
        if not os.path.exists(local_path_cfg):
            raise FileNotFoundError(f"Archivo local configurado no encontrado: '{local_path_cfg}'.")
        filename = local_path_cfg
        source_type = 'local_csv' if local_path_cfg.lower().endswith('.csv') else 'local_xlsx'
        logging.info(
            f"Se usará archivo local para '{param_from_frontend_url}': {filename}  (source_type ajustado a {source_type})"
        )

    logging.info(
        f"Iniciando carga de '{filename}' (Display Name: {param_from_frontend_url}, Tipo: {source_type})..."
    )

    # Usar selected_columns_check determinado al inicio (API o config)
    selected_columns = selected_columns_check
    if selected_columns:
        logging.info(f"Selección de columnas activa para '{param_from_frontend_url}': {len(selected_columns)} columnas - {selected_columns}")

    df_loaded = pd.DataFrame()

    try:
        if source_type == 'azure':
            connection_string = config_data.get("connection_string")
            container_name = config_data.get("container_name")
            if not all([connection_string, container_name, filename]):
                raise ValueError("Falta la configuración de Azure (ConnectionString, ContainerName, o Filename).")
            
            # Descarga optimizada con chunks y progreso
            blob_content = _download_blob_with_progress(connection_string, container_name, filename)

            if filename.lower().endswith('.csv'):
                df_loaded = csv_utils.read_csv_from_bytes(blob_content, filename, usecols=selected_columns)
            elif filename.lower().endswith(('.xlsx', '.xls')):
                if selected_columns:
                    df_loaded = pd.read_excel(io.BytesIO(blob_content), engine='openpyxl', usecols=selected_columns)
                else:
                    df_loaded = pd.read_excel(io.BytesIO(blob_content), engine='openpyxl')

        elif source_type == 'sharepoint':
            # Implementación de SharePoint con soporte para selección de columnas
            df_loaded = _read_file_from_sharepoint(filename, found_blob_attrs, usecols=selected_columns)

        elif source_type in ['local_xlsx', 'local_csv']:
            if not filename or not os.path.exists(filename):
                raise FileNotFoundError(f"Archivo local no encontrado: '{filename}'. Configúrelo en el lanzador.")
            if source_type == 'local_xlsx':
                if selected_columns:
                    df_loaded = pd.read_excel(filename, engine='openpyxl', usecols=selected_columns)
                else:
                    df_loaded = pd.read_excel(filename, engine='openpyxl')
            else:
                with open(filename, 'rb') as f:
                    file_bytes = f.read()
                df_loaded = csv_utils.read_csv_from_bytes(file_bytes, os.path.basename(filename), usecols=selected_columns)

        elif source_type == 'local_partitioned_csv':
            # Nuevo tipo de fuente: CSV particionados en directorio local
            logging.info(f"Cargando archivos CSV particionados desde: {filename}")

            # filename contiene el directorio base
            base_directory = filename

            # Obtener patrón de archivo desde configuración (default: *.csv)
            file_pattern = found_blob_attrs.get("file_pattern", "*.csv")

            # Validar que el directorio existe
            if not os.path.exists(base_directory):
                raise FileNotFoundError(
                    f"Directorio de particiones no encontrado: '{base_directory}'. "
                    f"Verifique que la ruta sea correcta y que OneDrive esté sincronizado."
                )

            # Progreso: descubrimiento
            progress_tracker.update_progress(
                10, "discovery",
                f"Descubriendo archivos con patrón '{file_pattern}'..."
            )

            # Cargar datos particionados utilizando la nueva función
            df_loaded = csv_utils.read_partitioned_csv_from_directory(
                base_directory=base_directory,
                file_pattern=file_pattern,
                usecols=selected_columns,  # Optimización de RAM
                log_prefix=param_from_frontend_url
            )

            progress_tracker.update_progress(
                25, "loaded",
                f"✓ Cargadas {len(df_loaded):,} filas desde particiones"
            )

        else:
            raise ValueError(f"Tipo de fuente ('_source_type') desconocido: '{source_type}'")

        if df_loaded.empty:
            logging.warning(f"El DataFrame para '{param_from_frontend_url}' está vacío después de la carga.")
            # Aún así, inicializar para evitar errores posteriores
            df_original = pd.DataFrame()
            current_blob_display_name = param_from_frontend_url
            return { "message": "Archivo cargado pero vacío.", "row_count_original": 0, "columns": [], "filter_options": {}, "source_type": source_type }
        
        # --- PROCESAMIENTO POST-CARGA ---
        df_loaded.columns = df_loaded.columns.str.strip().str.lower()
        logging.info(f"Cargadas {df_loaded.shape[0]} filas y {df_loaded.shape[1]} columnas.")
        progress_tracker.update_progress(25, "processing", f"Datos cargados: {df_loaded.shape[0]:,} filas. Procesando...")

        # --- FILTRADO PREVIO (PRE-FILTER) ---
        prefilter_column = found_blob_attrs.get("prefilter_nom_estado")
        if prefilter_column:
            logging.info(f"Aplicando filtrado previo: columna 'nom_estado' = '{prefilter_column}'")
            if 'nom_estado' in df_loaded.columns:
                df_loaded = df_loaded[df_loaded['nom_estado'] == prefilter_column]
                logging.info(f"Después del filtrado previo: {len(df_loaded)} filas restantes")
            else:
                logging.warning(f"Columna 'nom_estado' no encontrada para filtrado previo")

        progress_tracker.update_progress(40, "filtering", "Aplicando filtros y normalizaciones...")

        # --- ENRIQUECIMIENTO DE DATOS (DATA ENRICHMENT) ---
        enrichment_source = found_blob_attrs.get("enrichment_source")
        enrichment_join_column = found_blob_attrs.get("enrichment_join_column")
        enrichment_columns_str = found_blob_attrs.get("enrichment_columns")
        
        if enrichment_source and enrichment_join_column and enrichment_columns_str:
            logging.info(f"Iniciando enriquecimiento de datos con fuente: '{enrichment_source}'")
            progress_tracker.update_progress(55, "enrichment", f"Enriqueciendo datos desde '{enrichment_source}'...")

            try:
                # Cargar datos de enriquecimiento (con caché - HITO 1.2)
                df_enrichment = _load_enrichment_data_cached(enrichment_source, source_type, found_blob_attrs)
                
                if not df_enrichment.empty:
                    # Procesar columnas de enriquecimiento
                    enrichment_columns = [col.strip() for col in enrichment_columns_str.split(',') if col.strip()]
                    
                    # Verificar que las columnas existan en los datos de enriquecimiento
                    available_columns = [col for col in enrichment_columns if col in df_enrichment.columns]
                    if available_columns:
                        logging.info(f"Columnas de enriquecimiento disponibles: {available_columns}")
                        
                        # Realizar el JOIN
                        df_loaded = dataframe_utils._enrich_dataframe(df_loaded, df_enrichment, enrichment_join_column, available_columns)
                        logging.info(f"Enriquecimiento completado. Filas después del JOIN: {len(df_loaded)}")
                    else:
                        logging.warning(f"Ninguna de las columnas de enriquecimiento {enrichment_columns} encontrada en {enrichment_source}")
                else:
                    logging.warning(f"No se pudieron cargar datos de enriquecimiento desde '{enrichment_source}'")
                    
            except Exception as e:
                logging.error(f"Error durante el enriquecimiento de datos: {e}", exc_info=True)
                # Continuar sin enriquecimiento en caso de error

        # Normalizar valores de columnas SKU para evitar floats o espacios
        for _sku_col in ['ean_hijo', 'sku_hijo', 'sku_hijo_largo']:
            if _sku_col in df_loaded.columns:
                df_loaded[_sku_col] = (
                    df_loaded[_sku_col]
                    .astype(str)
                    .str.replace(r"\.0$", "", regex=True)
                    .str.strip()
                )

        current_config_blob_settings = config_data["filter_configs"].get(key_for_blob_options_lookup)
        if not current_config_blob_settings:
            raise ValueError(f"Configuración de filtros no encontrada para: {key_for_blob_options_lookup}")
            
        not_empty_cols_cfg = [col.lower() for col in current_config_blob_settings.get('not_empty_cols', [])]
        if not_empty_cols_cfg:
            df_loaded.dropna(subset=[col for col in not_empty_cols_cfg if col in df_loaded.columns], inplace=True)
        
        row_exclude_rules_cfg = {k.lower(): v for k,v in current_config_blob_settings.get('exclude_rows', {}).items()}
        if row_exclude_rules_cfg:
            for col, values in row_exclude_rules_cfg.items():
                if col in df_loaded.columns:
                    df_loaded = df_loaded[~df_loaded[col].isin(values)]

        # OPTIMIZACIÓN: Se elimina el paso de _optimize_dataframe_dtypes() porque
        # inmediatamente después se convierte todo a string, haciendo la optimización innecesaria.
        # Esto ahorra 25-45 segundos de procesamiento y reduce picos de memoria en equipos lentos.
        # Ver: OPTIMIZACION_RENDIMIENTO.md - Fase 1
        df_original = df_loaded
        current_blob_display_name = param_from_frontend_url

        # --- CONVERSIÓN SELECTIVA A STRING (OPTIMIZACIÓN FASE 1 - HITO 1.1) ---
        # Solo convertir columnas que realmente necesitan ser string:
        # 1. Columnas de filtro (definidas en config.ini)
        # 2. Columnas de display (si están especificadas)
        # 3. Columnas de SKU (sku_hijo, sku_padre, etc.)
        # Esto reduce el tiempo de conversión de 8-12s a 2-3s y ahorra ~60 MB RAM

        # Obtener columnas que necesitan conversión
        string_required_cols = set()

        # Agregar columnas de filtro (ya disponibles en cfg_filter_cols_list más adelante)
        cfg_filter_cols_list_preview = [col.lower().strip() for col in current_config_blob_settings.get('filter_cols', [])]
        if cfg_filter_cols_list_preview:
            string_required_cols.update(cfg_filter_cols_list_preview)

        # Agregar columnas de display si existen
        if selected_columns_from_api:
            string_required_cols.update([col.lower() for col in selected_columns_from_api])

        # Agregar columnas SKU conocidas (siempre deben ser string para joins)
        sku_columns = ['sku_hijo', 'sku_padre', 'sku_hijo_largo', 'sku_padre_largo',
                       'sku_padre_corto', 'cod_padre', 'ean', 'codigo', 'codigo_producto']
        string_required_cols.update([col for col in sku_columns])

        # Convertir solo las columnas necesarias que existen en el DataFrame
        existing_cols_to_convert = [col for col in string_required_cols if col in df_original.columns]

        if existing_cols_to_convert:
            logging.info(f"⚡ [OPTIMIZACIÓN] Conversión selectiva: {len(existing_cols_to_convert)} de {len(df_original.columns)} columnas a string")
            logging.info(f"⚡ [DEBUG] Columnas a convertir: {sorted(existing_cols_to_convert)}")

            conversion_errors = []
            for col in existing_cols_to_convert:
                try:
                    df_original[col] = df_original[col].astype(str)
                except Exception as e:
                    error_msg = f"Columna '{col}': {e}"
                    conversion_errors.append(error_msg)
                    logging.warning(f"⚠️ No se pudo convertir columna '{col}' a string: {e}")

            if conversion_errors:
                logging.error(f"❌ {len(conversion_errors)} columnas fallaron en conversión a string: {conversion_errors}")
            else:
                logging.info(f"✅ Todas las columnas convertidas exitosamente")
        else:
            logging.warning("⚠️ No se encontraron columnas para convertir a string - verificar configuración")

        # Limpiar strings de fechas con formatos problemáticos (doble slash, etc.)
        try:
            df_original = dataframe_utils.clean_date_strings(df_original)
            logging.info("✅ Limpieza de fechas aplicada exitosamente")
        except Exception as e:
            logging.warning(f"⚠️ Error durante limpieza de fechas: {e} - continuando sin limpieza")

        # Configurar DuckDB solo si está disponible
        if DUCKDB_AVAILABLE:
            if duckdb_conn:
                try: duckdb_conn.close()
                except Exception: pass
            duckdb_conn = duckdb.connect(database=':memory:')
            duckdb_conn.register('pandas_df', df_original)
            duckdb_conn.execute("CREATE OR REPLACE TABLE data AS SELECT * FROM pandas_df")
            duckdb_conn.unregister('pandas_df')
            logging.info(f"DuckDB configurado con {len(df_original)} filas")
        else:
            logging.info(f"Modo legacy - DuckDB no disponible, usando Pandas puro")

        logging.info(f"Procesamiento final completado. {len(df_original)} filas restantes.")
        progress_tracker.update_progress(75, "finalizing", "Procesamiento completado. Generando opciones de filtro...")

        # --- GENERACIÓN DE OPCIONES DE FILTRO (PARALELA - HITO 1.3) ---
        filter_options_api = {}
        MAX_FILTER_OPTIONS = 5000
        cfg_filter_cols_list = [col.lower().strip() for col in current_config_blob_settings.get('filter_cols', [])]
        cfg_hide_values_dict = {k.lower(): v for k, v in current_config_blob_settings.get('hide_values', {}).items()}

        # Función auxiliar para generar opciones de una columna (thread-safe)
        def _generate_filter_options_for_column(df: pd.DataFrame, col_name: str,
                                                 hide_values: set, max_options: int) -> tuple:
            """
            Genera opciones de filtro para una columna específica (thread-safe).

            Returns:
                Tupla (col_name, list_of_unique_values) o (col_name, None) si hay error/excede límite
            """
            try:
                if col_name not in df.columns:
                    logging.warning(f"⚠️ Columna de filtro '{col_name}' no existe en DataFrame")
                    return (col_name, None)

                # Obtener tipo de dato actual
                col_dtype = df[col_name].dtype
                logging.debug(f"Procesando filtro para columna '{col_name}' (tipo: {col_dtype})")

                # Convertir a string de forma segura
                try:
                    unique_vals = df[col_name].dropna().astype(str).unique()
                except Exception as conv_err:
                    logging.error(f"❌ Error convirtiendo columna '{col_name}' (tipo {col_dtype}) a string: {conv_err}")
                    # Intentar sin dropna como fallback
                    try:
                        unique_vals = df[col_name].astype(str).unique()
                        logging.warning(f"⚠️ Fallback exitoso para '{col_name}' sin dropna")
                    except Exception as fallback_err:
                        logging.error(f"❌ Fallback también falló para '{col_name}': {fallback_err}")
                        return (col_name, None)

                if len(unique_vals) > max_options:
                    logging.warning(f"Columna '{col_name}' excede el límite de {max_options} opciones de filtro ({len(unique_vals)} valores).")
                    return (col_name, None)

                # Filtrar valores ocultos y ordenar
                filtered_vals = sorted([val for val in unique_vals if val and val != 'nan' and val not in hide_values])

                logging.debug(f"✅ Columna '{col_name}': {len(filtered_vals)} opciones generadas")
                return (col_name, filtered_vals)

            except Exception as e:
                logging.error(f"❌ Error crítico generando opciones para columna '{col_name}': {e}", exc_info=True)
                return (col_name, None)

        # Generar opciones en paralelo usando ThreadPoolExecutor
        if cfg_filter_cols_list:
            import concurrent.futures

            logging.info(f"⚡ [OPTIMIZACIÓN] Generando opciones de filtro para {len(cfg_filter_cols_list)} columnas en paralelo...")
            logging.info(f"⚡ [DEBUG] Columnas de filtro: {cfg_filter_cols_list}")

            failed_columns = []
            successful_columns = []

            with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
                # Enviar todas las columnas al thread pool
                future_to_col = {
                    executor.submit(
                        _generate_filter_options_for_column,
                        df_original,
                        col,
                        set(cfg_hide_values_dict.get(col, [])),
                        MAX_FILTER_OPTIONS
                    ): col
                    for col in cfg_filter_cols_list
                }

                # Recoger resultados a medida que completan
                for future in concurrent.futures.as_completed(future_to_col):
                    try:
                        col_name, unique_values = future.result()
                        if unique_values is not None:  # Solo agregar si hay valores válidos
                            filter_options_api[col_name] = unique_values
                            successful_columns.append(col_name)
                        else:
                            failed_columns.append(col_name)
                    except Exception as e:
                        col_name = future_to_col.get(future, 'unknown')
                        failed_columns.append(col_name)
                        logging.error(f"❌ Excepción en thread pool para columna '{col_name}': {e}", exc_info=True)

            logging.info(f"✅ Opciones de filtro generadas: {len(filter_options_api)}/{len(cfg_filter_cols_list)} columnas exitosas")
            if failed_columns:
                logging.warning(f"⚠️ Columnas que fallaron en generación de filtros: {failed_columns}")
            if successful_columns:
                logging.info(f"✅ Columnas exitosas: {successful_columns}")
        else:
            logging.info("No hay columnas de filtro configuradas")

        # --- GUARDADO EN CACHÉ PERSISTENTE ---
        all_columns_list = list(df_original.columns)
        progress_tracker.update_progress(90, "caching", "Guardando datos en caché persistente...")

        # Guardar en cache persistente si es una base cacheable
        blob_config = get_blob_config_wrapper(param_from_frontend_url)
        source_url = blob_config.get('value', '') if blob_config else ''

        if persistent_cache.is_cacheable(param_from_frontend_url):
            success = persistent_cache.save_to_cache(param_from_frontend_url, df_original, source_url)
            if success:
                logging.info(f"✅ Datos de '{param_from_frontend_url}' guardados en cache persistente")
            else:
                logging.warning(f"⚠️ Fallo al guardar '{param_from_frontend_url}' en cache persistente")
        
        # Métricas de rendimiento
        load_time = time.time() - start_time
        logging.info(f"🚀 CARGA COMPLETADA: '{param_from_frontend_url}' en {load_time:.2f}s")
        logging.info(f"📊 ESTADÍSTICAS: {len(df_original):,} filas, {len(all_columns_list)} columnas, {source_type}")

        # Guardar preferencia de columnas si se usaron columnas específicas desde API
        if selected_columns_from_api:
            save_column_preference(param_from_frontend_url, selected_columns_from_api)

        # Reportar finalización exitosa al frontend
        progress_tracker.finish(
            success=True,
            final_message=f"✅ Carga completada: {len(df_original):,} filas en {load_time:.2f}s"
        )

        return {
            "message": f"Datos de '{filename}' cargados en {load_time:.2f}s.",
            "row_count_original": len(df_original),
            "columns": all_columns_list,
            "filter_options": filter_options_api,
            "source_type": source_type,
            "from_cache": False,
            "load_time_seconds": round(load_time, 2),
            "cache_decision": cache_decision if cache_decision == "downloading_fresh" else "no_cache"
        }

    except Exception as e:
        logging.error(f"Error crítico durante la carga de '{param_from_frontend_url}': {e}", exc_info=True)

        # Reportar error al frontend
        progress_tracker.error(f"❌ Error en la carga: {str(e)}")

        df_original = pd.DataFrame()
        current_blob_display_name = None
        raise

def refresh_blob_data(param_from_frontend_url: str, selected_columns_from_api: Optional[List[str]] = None) -> Dict[str, Any]:
    """
    Fuerza la recarga de datos desde la fuente original, actualizando el cache persistente.

    Args:
        param_from_frontend_url: Identificador de la fuente de datos
        selected_columns_from_api: Lista opcional de columnas específicas desde API
    """
    logging.info(f"refresh_blob_data: Solicitud para forzar actualización de '{param_from_frontend_url}'.")

    # Limpiar cache persistente para forzar descarga completa
    if persistent_cache.is_cacheable(param_from_frontend_url):
        if persistent_cache.has_cached_data(param_from_frontend_url):
            success = persistent_cache.clear_cache(param_from_frontend_url)
            if success:
                logging.info(f"✅ Cache persistente eliminado para '{param_from_frontend_url}' - se forzará descarga completa")
            else:
                logging.warning(f"⚠️ No se pudo limpiar el cache persistente para '{param_from_frontend_url}'")

    # Llamar a load_blob_data; como la caché está vacía para esta clave,
    # se recargará desde la fuente y se repoblará la caché.
    # Los filtros SKU globales se limpiarán dentro de load_blob_data.
    return load_blob_data(param_from_frontend_url, selected_columns_from_api=selected_columns_from_api)




# Smart cache cleanup function eliminated - persistent cache handles all caching


# Funciones de filtrado
def _clear_filter_states(use_sku_hijo_file: bool, sku_hijo_manual_list: Optional[List[str]],
                        use_sku_padre_file: bool, sku_padre_manual_list: Optional[List[str]],
                        use_ticket_file: bool, ticket_manual_list: Optional[List[str]]):
    """Limpia estados globales de filtros según condiciones."""
    global sku_hijo_filter_list, sku_padre_filter_list, ticket_filter_list
    
    if not use_sku_hijo_file or sku_hijo_manual_list:
        sku_hijo_filter_list = None
        logging.info("Estado 'sku_hijo_filter_list' limpiado")
    
    if not use_sku_padre_file or sku_padre_manual_list:
        sku_padre_filter_list = None
        logging.info("Estado 'sku_padre_filter_list' limpiado")
    
    if not use_ticket_file or ticket_manual_list:
        ticket_filter_list = None
        logging.info("Estado 'ticket_filter_list' limpiado")

def _log_filter_diagnostics():
    """Registra información de diagnóstico para filtros."""
    logging.debug("=== DIAGNÓSTICO apply_all_filters ===")
    logging.debug(f"current_blob_display_name: {current_blob_display_name}")
    logging.debug(f"df_original.shape: {df_original.shape if not df_original.empty else 'N/A'}")
    logging.debug("=== FIN DIAGNÓSTICO ===")



def apply_all_filters(
    value_filters: Dict[str, List[str]],
    use_sku_hijo_file: bool,
    extend_sku_hijo: bool,
    sku_hijo_manual_list: Optional[List[str]],
    use_sku_padre_file: bool,
    sku_padre_manual_list: Optional[List[str]],
    use_ticket_file: bool,
    ticket_manual_list: Optional[List[str]],
    lineamiento_manual_list: Optional[List[str]],
    selected_display_columns: Optional[List[str]] = None,
    enable_priority_coloring: bool = False,
    page: int = 1,
    page_size: int = 100,
    custom_text_filters: Optional[Dict[str, List[str]]] = None
) -> Dict[str, Any]:
    """Aplica todos los filtros a los datos."""
    global df_filtered, df_original, duckdb_conn, current_blob_display_name, config_data
    
    # Limpiar estados de filtros
    _clear_filter_states(use_sku_hijo_file, sku_hijo_manual_list, 
                        use_sku_padre_file, sku_padre_manual_list,
                        use_ticket_file, ticket_manual_list)
    
    # Inicializar listas de no encontrados
    skus_no_encontrados_hijo = []
    skus_no_encontrados_padre = []
    tickets_no_encontrados = []
    lineamientos_no_encontrados = []
    
    # Diagnóstico
    _log_filter_diagnostics()
    
    # Validaciones de datos vacíos
    if df_original.empty and not df_original.columns.any():
        return _get_empty_filter_response()
    
    if df_original.empty and df_original.columns.any():
        output_cols = ([col for col in selected_display_columns if col in df_original.columns] 
                      if selected_display_columns else list(df_original.columns))
        return _get_empty_filter_response(output_cols)
    
    logging.info(f"Aplicando filtros via DuckDB. Filas iniciales: {len(df_original)}")
    
    # Obtener columnas SKU candidatas
    current_source_type = config_data["blob_options"][current_blob_display_name.upper()].get("source_type") if current_blob_display_name else None
    sku_hijo_candidates, sku_padre_candidates = _get_sku_column_candidates(current_source_type, current_blob_display_name)
    
    # Buscar la primera columna existente para cada filtro
    sku_col_hijo_to_use = dataframe_utils.find_first_existing_column(df_original, sku_hijo_candidates)
    sku_col_padre_to_use = dataframe_utils.find_first_existing_column(df_original, sku_padre_candidates)
    
    # Buscar columna color (maneja diferentes variaciones como 'color', 'COLOR', etc.)
    color_candidates = ['color', 'COLOR', 'Color']
    color_col_to_use = dataframe_utils.find_first_existing_column(df_original, color_candidates)

    logging.info(f"Columna SKU hijo seleccionada: {sku_col_hijo_to_use}")
    logging.info(f"Columna SKU padre seleccionada: {sku_col_padre_to_use}")
    logging.info(f"Columna color seleccionada: {color_col_to_use}")

    # --- FIN DE SELECCIÓN ROBUSTA DE COLUMNAS SKU ---

    # --- NUEVO: Buscar variantes de 'ean_hijo' si no existe la columna ---
    # (Este bloque ya no es necesario porque find_first_existing_column ya busca variantes)

    def _quote(val: str) -> str:
        return val.replace("'", "''")

    where_clauses = []
    
    # --- AÑADIR ESTE BLOQUE DE DEBUG ---
    logging.info("--- DEBUG apply_all_filters ---")
    logging.info(f"use_sku_hijo_file: {use_sku_hijo_file}")
    logging.info(f"Contenido de sku_hijo_filter_list: {sku_hijo_filter_list is not None}")
    logging.info(f"use_sku_padre_file: {use_sku_padre_file}")
    logging.info(f"Contenido de sku_padre_filter_list: {sku_padre_filter_list is not None}")
    logging.info(f"sku_col_hijo_to_use: {sku_col_hijo_to_use}")
    logging.info(f"sku_col_padre_to_use: {sku_col_padre_to_use}")
    logging.info(f"Columnas disponibles en df_original: {list(df_original.columns)}")
    logging.info("-----------------------------")
    # --- FIN DEL BLOQUE DE DEBUG ---

    for col_name, selected_values in (value_filters or {}).items():
        # Buscar la columna que coincida (case-insensitive)
        matching_col = None
        for col in df_original.columns:
            if col.lower() == col_name.lower():
                matching_col = col
                break
        
        if matching_col and selected_values:
            actual_vals = ['' if v == "[Vacío]" else str(v) for v in selected_values]
            val_list = ','.join(f"'{_quote(v)}'" for v in actual_vals)
            # Usar el nombre real de la columna (con su capitalización original)
            where_clauses.append(f"coalesce(trim(\"{matching_col}\"), '') in ({val_list})")
            logging.info(f"Aplicando filtro a columna '{matching_col}' con valores: {selected_values}")

    # --- INICIO DE LA CORRECCIÓN DEL FILTRO SKU ---

    # Filtro SKU Hijo
    skus_hijo_a_filtrar = set()
    # Añadir SKUs del archivo SOLO si el frontend lo indica
    if use_sku_hijo_file and sku_hijo_filter_list:
        skus_hijo_a_filtrar.update(str(s).strip() for s in sku_hijo_filter_list if str(s).strip())
    # Añadir SIEMPRE los SKUs de la lista manual si existen
    if sku_hijo_manual_list:
        skus_hijo_a_filtrar.update(str(s).strip() for s in sku_hijo_manual_list if str(s).strip())

    if skus_hijo_a_filtrar:
        if sku_col_hijo_to_use and sku_col_hijo_to_use in df_original.columns:
            val_list_hijo = ','.join(f"'{_quote(v)}'" for v in skus_hijo_a_filtrar)
            if extend_sku_hijo and sku_col_padre_to_use and color_col_to_use:
                # Usar DuckDB si está disponible, sino fallback a Pandas
                if DUCKDB_AVAILABLE:
                    # Verificar que DuckDB esté inicializado
                    if duckdb_conn is None:
                        logging.error("DuckDB connection es None durante extend_sku_hijo - reinicializando conexión")
                        _setup_duckdb_connection(df_original)

                    pair_query = (
                        f"SELECT DISTINCT trim(CAST({sku_col_padre_to_use} AS VARCHAR)) as padre, trim({color_col_to_use}) as color "
                        f"FROM data WHERE trim(CAST({sku_col_hijo_to_use} AS VARCHAR)) in ({val_list_hijo})"
                    )
                    pair_df = duckdb_conn.execute(pair_query).fetchdf()
                else:
                    # Fallback a Pandas puro (modo legacy)
                    logging.info("Usando Pandas puro para extend_sku_hijo (modo legacy)")
                    mask = df_original[sku_col_hijo_to_use].astype(str).str.strip().isin(skus_hijo_a_filtrar)
                    pair_df = df_original.loc[mask, [sku_col_padre_to_use, color_col_to_use]].copy()
                    pair_df.columns = ['padre', 'color']
                    pair_df['padre'] = pair_df['padre'].astype(str).str.strip()
                    pair_df['color'] = pair_df['color'].astype(str).str.strip()
                    pair_df = pair_df.drop_duplicates()

                if not pair_df.empty:
                    # Separar pares con color válido de pares sin color
                    # Consideramos sin color: NULL, vacío, 'nan', 'None'
                    pairs_with_color = []
                    padres_without_color = []

                    for _, row in pair_df.iterrows():
                        padre_val = str(row['padre']).strip()
                        color_val = str(row['color']).strip()

                        # Verificar si el color es válido (no NULL, no vacío, no 'nan', no 'None')
                        if color_val and color_val.lower() not in ('nan', 'none', 'null', ''):
                            pairs_with_color.append((padre_val, color_val))
                        else:
                            padres_without_color.append(padre_val)

                    # Construir WHERE clauses separadas
                    where_parts = []

                    # Parte 1: Pares con color (padre, color)
                    if pairs_with_color:
                        sku_extension_pairs = [f"('{_quote(p)}','{_quote(c)}')" for p, c in pairs_with_color]
                        where_parts.append(
                            f"(trim(CAST({sku_col_padre_to_use} AS VARCHAR)), trim({color_col_to_use})) in ({','.join(sku_extension_pairs)})"
                        )

                    # Parte 2: Padres sin color (solo padre, color NULL/vacío)
                    if padres_without_color:
                        padres_list = ','.join(f"'{_quote(p)}'" for p in padres_without_color)
                        where_parts.append(
                            f"(trim(CAST({sku_col_padre_to_use} AS VARCHAR)) in ({padres_list}) AND "
                            f"(trim({color_col_to_use}) IS NULL OR trim({color_col_to_use}) = '' OR trim({color_col_to_use}) = 'nan' OR trim({color_col_to_use}) = 'None'))"
                        )

                    # Combinar ambas partes con OR
                    if where_parts:
                        if len(where_parts) == 1:
                            where_clauses.append(where_parts[0])
                        else:
                            where_clauses.append(f"({' OR '.join(where_parts)})")
                    else:
                        # Si no hay pares válidos, fallar la consulta
                        where_clauses.append("1=0")
                else:
                    # Si no se encuentran pares, la consulta debe fallar para no devolver todo el DataFrame
                    where_clauses.append("1=0") 
            else:
                where_clauses.append(f"trim(CAST({sku_col_hijo_to_use} AS VARCHAR)) in ({val_list_hijo})")
        else:
            logging.warning("No se encontró ninguna columna válida para el filtro SKU hijo.")
            # Si no hay columna, todos los SKUs buscados se consideran no encontrados.
            skus_no_encontrados_hijo = list(skus_hijo_a_filtrar)

    # Filtro SKU Padre
    skus_padre_a_filtrar = set()
    # Añadir SKUs del archivo SOLO si el frontend lo indica
    if use_sku_padre_file and sku_padre_filter_list:
        skus_padre_a_filtrar.update(str(s).strip() for s in sku_padre_filter_list if str(s).strip())
    # Añadir SIEMPRE los SKUs de la lista manual si existen
    if sku_padre_manual_list:
        skus_padre_a_filtrar.update(str(s).strip() for s in sku_padre_manual_list if str(s).strip())

    logging.info(f"SKUs padre a filtrar: {skus_padre_a_filtrar}")
    logging.info(f"Columna SKU padre a usar: {sku_col_padre_to_use}")
    logging.info(f"¿Columna existe en df_original? {sku_col_padre_to_use in df_original.columns if sku_col_padre_to_use else False}")

    if skus_padre_a_filtrar:
        if sku_col_padre_to_use and sku_col_padre_to_use in df_original.columns:
            val_list_padre = ','.join(f"'{_quote(v)}'" for v in skus_padre_a_filtrar)
            where_clause_padre = f"trim(CAST({sku_col_padre_to_use} AS VARCHAR)) in ({val_list_padre})"
            where_clauses.append(where_clause_padre)
            logging.info(f"Filtro SKU padre aplicado: {where_clause_padre}")
        else:
            logging.warning("No se encontró ninguna columna válida para el filtro SKU padre.")
            # Si no hay columna, todos los SKUs buscados se consideran no encontrados.
            skus_no_encontrados_padre = list(skus_padre_a_filtrar)

# --- FIN DE LA CORRECCIÓN DEL FILTRO SKU ---

    # ... (el resto de la función continúa sin cambios)
    tickets_set = set()
    if use_ticket_file and ticket_filter_list:
        tickets_set.update(str(t).strip().lower() for t in ticket_filter_list if str(t).strip())
    if ticket_manual_list:
        tickets_set.update(str(t).strip().lower() for t in ticket_manual_list if str(t).strip())

    if tickets_set:
        if 'ticket' in df_original.columns:
            val_list = ','.join(f"'{_quote(v)}'" for v in tickets_set)
            where_clauses.append(f"lower(coalesce(ticket,'')) in ({val_list})")
        else:
            logging.warning("Columna 'ticket' no encontrada para el filtro por ticket.")

    if lineamiento_manual_list:
        terms = [t.strip() for t in lineamiento_manual_list if t.strip()]
        if terms:
            if 'asunto_lineamientos' in df_original.columns:
                like_clauses = [f"lower(asunto_lineamientos) like '%{_quote(t.lower())}%'" for t in terms]
                where_clauses.append('(' + ' OR '.join(like_clauses) + ')')
            else:
                logging.warning("Columna 'asunto_lineamientos' no encontrada para filtro lineamiento.")

    # Filtros personalizados de texto (búsqueda por coincidencias)
    if custom_text_filters:
        for column_name, search_terms in custom_text_filters.items():
            if not search_terms:
                continue

            # Limpiar y validar términos
            terms = [t.strip() for t in search_terms if t.strip()]
            if not terms:
                continue

            # Verificar que la columna exista (case-insensitive)
            matching_col = None
            for col in df_original.columns:
                if col.lower() == column_name.lower():
                    matching_col = col
                    break

            if matching_col:
                # Crear cláusulas de igualdad exacta (case-insensitive)
                exact_clauses = [
                    f"lower(CAST(\"{matching_col}\" AS VARCHAR)) = '{_quote(t.lower())}'"
                    for t in terms
                ]
                where_clauses.append('(' + ' OR '.join(exact_clauses) + ')')
                logging.info(f"Filtro personalizado aplicado en columna '{matching_col}': {len(terms)} términos (búsqueda exacta)")
            else:
                logging.warning(
                    f"Columna '{column_name}' no encontrada para filtro personalizado - ignorando"
                )

    # Ejecutar filtrado usando DuckDB o Pandas según disponibilidad
    if DUCKDB_AVAILABLE:
        # Modo completo con DuckDB
        query = "SELECT * FROM data"
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        # CRÍTICO: Verificar que DuckDB esté inicializado antes de usarlo
        if duckdb_conn is None:
            logging.error("DuckDB connection es None durante filtrado - reinicializando conexión")
            _setup_duckdb_connection(df_original)

            if duckdb_conn is None:
                logging.error("No se pudo reinicializar DuckDB connection")
                raise Exception("DuckDB connection no disponible para filtrado")

        df_filtered = duckdb_conn.execute(query).fetchdf()
        df_filtered.reset_index(drop=True, inplace=True)
        logging.info(f"Filtrado ejecutado con DuckDB - {len(df_filtered)} filas resultantes")
    else:
        # Modo legacy con Pandas puro - más lento pero funciona sin DuckDB
        logging.warning("Usando modo legacy con Pandas puro - el filtrado será más lento")
        df_filtered = df_original.copy()

        # Aplicar filtros de columnas value_filters
        for col_name, selected_values in value_filters.items():
            if selected_values and col_name in df_filtered.columns:
                # Manejar valores vacíos
                actual_vals = ['' if v == "[Vacío]" else str(v) for v in selected_values]
                df_filtered = df_filtered[df_filtered[col_name].astype(str).str.strip().isin(actual_vals)]

        # Aplicar filtro SKU hijo (sin extensión en modo legacy)
        if skus_hijo_a_filtrar and sku_col_hijo_to_use and sku_col_hijo_to_use in df_filtered.columns:
            df_filtered = df_filtered[df_filtered[sku_col_hijo_to_use].astype(str).str.strip().isin(skus_hijo_a_filtrar)]

        # Aplicar filtro SKU padre
        if skus_padre_a_filtrar and sku_col_padre_to_use and sku_col_padre_to_use in df_filtered.columns:
            df_filtered = df_filtered[df_filtered[sku_col_padre_to_use].astype(str).str.strip().isin(skus_padre_a_filtrar)]

        # Aplicar filtro tickets
        if tickets_set and 'ticket' in df_filtered.columns:
            df_filtered = df_filtered[df_filtered['ticket'].astype(str).str.lower().str.strip().isin(tickets_set)]

        # Aplicar filtro lineamientos (búsqueda de texto)
        if lineamiento_manual_list and 'asunto_lineamientos' in df_filtered.columns:
            terms = [t.strip().lower() for t in lineamiento_manual_list if t.strip()]
            if terms:
                mask = df_filtered['asunto_lineamientos'].astype(str).str.lower().str.contains('|'.join(terms), na=False, regex=True)
                df_filtered = df_filtered[mask]

        df_filtered.reset_index(drop=True, inplace=True)
        logging.info(f"Filtrado ejecutado con Pandas (modo legacy) - {len(df_filtered)} filas resultantes")

    # --- CÁLCULO DE SKUS NO ENCONTRADOS ---
    if skus_hijo_a_filtrar:
        if sku_col_hijo_to_use and sku_col_hijo_to_use in df_filtered.columns:
            # Asegurarse que los tipos son consistentes para la comparación
            encontrados_hijo = set(df_filtered[sku_col_hijo_to_use].astype(str).str.strip().unique())
            skus_no_encontrados_hijo = sorted(list(skus_hijo_a_filtrar - encontrados_hijo))
            logging.info(f"{len(skus_no_encontrados_hijo)} SKUs hijos no encontrados.")
        elif not (sku_col_hijo_to_use and sku_col_hijo_to_use in df_original.columns):
            # Este caso ya se maneja arriba, pero lo mantenemos por claridad.
            skus_no_encontrados_hijo = sorted(list(skus_hijo_a_filtrar))

    if skus_padre_a_filtrar:
        if sku_col_padre_to_use and sku_col_padre_to_use in df_filtered.columns:
            encontrados_padre = set(df_filtered[sku_col_padre_to_use].astype(str).str.strip().unique())
            skus_no_encontrados_padre = sorted(list(skus_padre_a_filtrar - encontrados_padre))
            logging.info(f"{len(skus_no_encontrados_padre)} SKUs padres no encontrados.")
        elif not (sku_col_padre_to_use and sku_col_padre_to_use in df_original.columns):
            skus_no_encontrados_padre = sorted(list(skus_padre_a_filtrar))

    # --- CÁLCULO DE TICKETS NO ENCONTRADOS (REQUERIMIENTOS) ---
    if tickets_set:
        if 'ticket' in df_filtered.columns:
            encontrados_tickets = set(df_filtered['ticket'].astype(str).str.strip().str.lower().unique())
            # Filtrar valores vacíos del conjunto encontrado
            encontrados_tickets = {t for t in encontrados_tickets if t and t != 'nan'}
            tickets_no_encontrados = sorted(list(tickets_set - encontrados_tickets))
            logging.info(f"{len(tickets_no_encontrados)} Tickets (requerimientos) no encontrados.")
        elif 'ticket' not in df_original.columns:
            # Si no existe la columna, todos los tickets buscados no se encontraron
            tickets_no_encontrados = sorted(list(tickets_set))
            logging.warning(f"Columna 'ticket' no existe, todos los tickets se marcan como no encontrados: {len(tickets_no_encontrados)}")

    # --- CÁLCULO DE LINEAMIENTOS NO ENCONTRADOS (TICKETS) ---
    if lineamiento_manual_list:
        terms = [t.strip() for t in lineamiento_manual_list if t.strip()]
        if terms and 'asunto_lineamientos' in df_filtered.columns:
            # Para lineamientos usamos búsqueda parcial (LIKE), así que verificamos si alguno coincide
            encontrados_lineamientos = set()
            asuntos_en_resultados = df_filtered['asunto_lineamientos'].astype(str).str.lower().unique()

            for term in terms:
                term_lower = term.lower()
                # Verificar si el término aparece en algún asunto
                if any(term_lower in asunto for asunto in asuntos_en_resultados if asunto and asunto != 'nan'):
                    encontrados_lineamientos.add(term)

            lineamientos_no_encontrados = sorted([t for t in terms if t not in encontrados_lineamientos])
            logging.info(f"{len(lineamientos_no_encontrados)} Lineamientos (tickets) no encontrados.")
        elif 'asunto_lineamientos' not in df_original.columns:
            # Si no existe la columna, todos los lineamientos buscados no se encontraron
            lineamientos_no_encontrados = sorted([t.strip() for t in lineamiento_manual_list if t.strip()])
            logging.warning(f"Columna 'asunto_lineamientos' no existe, todos los lineamientos se marcan como no encontrados: {len(lineamientos_no_encontrados)}")

    output_columns_final = [col for col in selected_display_columns if col in df_filtered.columns] if selected_display_columns else list(df_filtered.columns)
    page = max(page, 1)
    page_size = max(page_size, 1)
    start_idx = (page - 1) * page_size
    end_idx = start_idx + page_size
    df_to_send = (
        df_filtered[output_columns_final].iloc[start_idx:end_idx]
        if not df_filtered.empty
        else pd.DataFrame(columns=output_columns_final)
    )
    
    # Limpiar valores NaN y NaT para mejor visualización en el frontend
    df_to_send = dataframe_utils.clean_nan_nat_values(df_to_send)

    # Siempre verificar si hay columna de prioridad disponible
    has_priority_col = dataframe_utils._has_priority_column(df_filtered)
    
    # Siempre procesar información de prioridad si hay columna disponible (para habilitar toggle dinámico)
    priority_info = {}
    if has_priority_col:
        # Usar df_filtered para detectar columna y obtener conteos totales, df_to_send para datos de la página
        priority_info = dataframe_utils._get_priority_info(df_filtered, df_to_send, start_idx)
        # Solo incluir en respuesta si el coloreado está habilitado o si es la primera carga
        logging.info(f"Columna de prioridad disponible - procesando información para toggle dinámico")

    result = {
        "row_count_filtered": len(df_filtered),
        "data": df_to_send.to_dict(orient='records'),
        "columns_in_data": output_columns_final,
        "page": page,
        "page_size": page_size,
        "skus_no_encontrados_hijo": skus_no_encontrados_hijo,
        "skus_no_encontrados_padre": skus_no_encontrados_padre,
        "tickets_no_encontrados": tickets_no_encontrados,
        "lineamientos_no_encontrados": lineamientos_no_encontrados,
        "has_priority_column": has_priority_col  # Siempre informar si hay columna disponible
    }
    
    # Agregar información de prioridad si hay columna disponible (siempre que exista para toggle dinámico)
    if has_priority_col and priority_info:
        result["priority_info"] = priority_info
        
    return result

# --- FUNCIÓN DE EXPORTACIÓN MODIFICADA CON CANCELACIÓN ---
def get_excel_export(
    value_filters: Dict[str, List[str]],
    use_sku_hijo_file: bool,
    extend_sku_hijo: bool,
    sku_hijo_manual_list: Optional[List[str]],
    use_sku_padre_file: bool,
    sku_padre_manual_list: Optional[List[str]],
    use_ticket_file: bool,
    ticket_manual_list: Optional[List[str]],
    lineamiento_manual_list: Optional[List[str]],
    selected_display_columns: Optional[List[str]] = None,
    enable_priority_coloring: bool = False,
    custom_text_filters: Optional[Dict[str, List[str]]] = None
) -> str:
    global df_filtered, df_original, current_export_task
    
    # Resetear estado de cancelación al inicio
    reset_export_cancellation()
    current_export_task = "excel_export"
    
    try:
        # Verificar cancelación antes de empezar
        check_export_cancellation()
        
        # Llama a la función de filtrado actualizada para obtener el DataFrame completo
        filter_results = apply_all_filters(
            value_filters,
            use_sku_hijo_file,
            extend_sku_hijo,
            sku_hijo_manual_list,
            use_sku_padre_file,
            sku_padre_manual_list,
            use_ticket_file,
            ticket_manual_list,
            lineamiento_manual_list,
            selected_display_columns,
            enable_priority_coloring,
            page=1,
            page_size=1,
            custom_text_filters=custom_text_filters
        )
        
        # Verificar cancelación después del filtrado
        check_export_cancellation()
        
        if df_filtered.empty and not df_original.columns.any():
            raise ValueError("No hay datos para exportar.")
            
        export_cols_final = filter_results["columns_in_data"]
        
        if not export_cols_final and df_original.columns.any():
            export_cols_final = list(df_original.columns)
        elif not export_cols_final:
            raise ValueError("No hay columnas definidas para exportar.")

        # Preparar DataFrame a exportar y limpiar NaN/Inf
        df_to_export = df_filtered[export_cols_final] if not df_filtered.empty else pd.DataFrame(columns=export_cols_final)
        
        # Limpiar valores NaN y NaT para mejor visualización en exportaciones
        df_to_export = dataframe_utils.clean_nan_nat_values(df_to_export)

        # Verificar cancelación antes de crear el archivo
        check_export_cancellation()

        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_file.close()

        workbook = xlsxwriter.Workbook(tmp_file.name, {
            'constant_memory': True,
            'in_memory': True,
            'nan_inf_to_errors': True  # Evita error con NaN/Inf
        })
        worksheet = workbook.add_worksheet('DatosFiltrados')

        # Definir formatos de prioridad si está habilitado (colores suavizados)
        priority_formats = {}
        if enable_priority_coloring:
            priority_formats = {
                'PRIORIDAD_1': workbook.add_format({'bg_color': '#FFCDD2', 'font_color': '#B71C1C'}),  # Rojo suave con texto rojo oscuro
                'PRIORIDAD_2': workbook.add_format({'bg_color': '#FFF9C4', 'font_color': '#F57F17'}),  # Amarillo suave con texto amarillo oscuro
                'PRIORIDAD_3': workbook.add_format({'bg_color': '#C8E6C9', 'font_color': '#1B5E20'})   # Verde suave con texto verde oscuro
            }

        # Escribir encabezados
        for col_num, col_name in enumerate(export_cols_final):
            worksheet.write(0, col_num, col_name)

        # Escribir datos con verificación de cancelación y coloreado por prioridad
        if not df_to_export.empty:
            total_rows = len(df_to_export)
            
            # Generar información de prioridad para TODO el DataFrame de exportación
            export_priority_info = {}
            if enable_priority_coloring and dataframe_utils._has_priority_column(df_filtered):
                export_priority_info = dataframe_utils._get_priority_info(df_filtered, df_to_export, 0)
                logging.info(f"Información de prioridad generada para {total_rows} filas de exportación")
            
            for row_idx, row in enumerate(df_to_export.itertuples(index=False, name=None), start=1):
                # Verificar cancelación cada 1000 filas
                if row_idx % 1000 == 0:
                    check_export_cancellation()
                    logging.info(f"Progreso de exportación Excel: {row_idx}/{total_rows} filas procesadas")

                # Aplicar coloreado por prioridad si está habilitado
                row_format = None
                if enable_priority_coloring and export_priority_info.get("has_priority_column", False):
                    # row_idx-1 porque enumerate empieza en 1 pero los índices de datos empiezan en 0
                    data_row_idx = row_idx - 1
                    priority_value = export_priority_info.get("row_priorities", {}).get(data_row_idx)

                    if priority_value:
                        # Hacer matching más flexible con los valores de prioridad
                        if any(p in priority_value.upper() for p in ["PRIORIDAD_1", "PRIORITY_1"]):
                            row_format = priority_formats.get('PRIORIDAD_1')
                        elif any(p in priority_value.upper() for p in ["PRIORIDAD_2", "PRIORITY_2"]):
                            row_format = priority_formats.get('PRIORIDAD_2')
                        elif any(p in priority_value.upper() for p in ["PRIORIDAD_3", "PRIORITY_3"]):
                            row_format = priority_formats.get('PRIORIDAD_3')

                        # Debug logging para las primeras 5 filas
                        if row_idx <= 5:
                            logging.info(f"Fila {row_idx}: priority_value='{priority_value}', row_format={'aplicado' if row_format else 'ninguno'}")

                # Convertir NaT y NaN a None para evitar errores en xlsxwriter
                # xlsxwriter no puede manejar pd.NaT directamente
                cleaned_row = []
                for value in row:
                    if pd.isna(value):  # Detecta tanto NaN como NaT
                        cleaned_row.append(None)
                    else:
                        cleaned_row.append(value)

                # Escribir fila con formato si aplica
                if row_format:
                    for col_idx, value in enumerate(cleaned_row):
                        worksheet.write(row_idx, col_idx, value, row_format)
                else:
                    worksheet.write_row(row_idx, 0, cleaned_row)

        workbook.close()

        logging.info(f"Exportando {len(df_to_export)} filas y {len(export_cols_final)} columnas a Excel.")
        return tmp_file.name
        
    except InterruptedError:
        # Limpiar archivo temporal si existe
        if 'tmp_file' in locals():
            try:
                os.unlink(tmp_file.name)
            except:
                pass
        raise
    finally:
        current_export_task = None


def get_csv_export(
    value_filters: Dict[str, List[str]],
    use_sku_hijo_file: bool,
    extend_sku_hijo: bool,
    sku_hijo_manual_list: Optional[List[str]],
    use_sku_padre_file: bool,
    sku_padre_manual_list: Optional[List[str]],
    use_ticket_file: bool,
    ticket_manual_list: Optional[List[str]],
    lineamiento_manual_list: Optional[List[str]],
    selected_display_columns: Optional[List[str]] = None,
    enable_priority_coloring: bool = False,
    custom_text_filters: Optional[Dict[str, List[str]]] = None
) -> Callable[[], Iterator[bytes]]:
    """Preparar un generador de bytes CSV con los datos filtrados."""
    global df_filtered, df_original, current_export_task
    
    # Resetear estado de cancelación al inicio
    reset_export_cancellation()
    current_export_task = "csv_export"
    
    try:
        # Verificar cancelación antes de empezar
        check_export_cancellation()
        
        filter_results = apply_all_filters(
            value_filters,
            use_sku_hijo_file,
            extend_sku_hijo,
            sku_hijo_manual_list,
            use_sku_padre_file,
            sku_padre_manual_list,
            use_ticket_file,
            ticket_manual_list,
            lineamiento_manual_list,
            selected_display_columns,
            enable_priority_coloring,
            custom_text_filters=custom_text_filters
        )
        
        # Verificar cancelación después del filtrado
        check_export_cancellation()

        if df_filtered.empty and not df_original.columns.any():
            raise ValueError("No hay datos para exportar.")

        export_cols_final = filter_results["columns_in_data"]

        if not export_cols_final and df_original.columns.any():
            export_cols_final = list(df_original.columns)
        elif not export_cols_final:
            raise ValueError("No hay columnas definidas para exportar.")

        df_to_export = df_filtered[export_cols_final] if not df_filtered.empty else pd.DataFrame(columns=export_cols_final)
        
        # Limpiar NaN y NaT para mejor visualización en CSV
        df_to_export = dataframe_utils.clean_nan_nat_values(df_to_export)

        def csv_generator(chunk_size: int = 1000) -> Iterator[bytes]:
            try:
                yield (",".join(export_cols_final) + "\n").encode("utf-8")
                total_rows = len(df_to_export)
                for start in range(0, total_rows, chunk_size):
                    # Verificar cancelación cada chunk
                    check_export_cancellation()
                    
                    chunk_df = df_to_export.iloc[start : start + chunk_size]
                    yield chunk_df.to_csv(index=False, header=False).encode("utf-8")
                    
                    # Log de progreso cada 10 chunks
                    if start % (chunk_size * 10) == 0:
                        logging.info(f"Progreso de exportación CSV: {start}/{total_rows} filas procesadas")
                        
            except InterruptedError:
                logging.info("Exportación CSV cancelada por el usuario.")
                raise

        logging.info(f"Exportando {len(df_to_export)} filas y {len(export_cols_final)} columnas a CSV.")
        return csv_generator
        
    except InterruptedError:
        raise
    finally:
        current_export_task = None


# (save/load_filter_state_from_config no cambian)
# ... resto de funciones sin cambios ...
def save_current_filter_state_to_config(
    blob_display_name: str, # original_display_name (case preservado)
    value_filters_state: Dict[str, List[str]],
    selected_columns_state: List[str],
    extend_sku_search_state: bool
) -> bool:
    # ... código sin cambios ...
    global config_data
    
    parser = config_data.get("config_parser")
    if not parser:
        logging.error("ConfigParser no está cargado. No se puede guardar el estado.")
        return False

    # La clave para la sección SavedState debe ser la versión .upper() del original_display_name
    key_upper_for_saved_state_section = blob_display_name.upper()
        
    section_name = f"SavedState_{key_upper_for_saved_state_section}"
    try:
        if not parser.has_section(section_name):
            parser.add_section(section_name)
        else: # Limpiar sección existente antes de guardar nuevos valores
            for key in parser.options(section_name):
                parser.remove_option(section_name, key)

        # Guardar filtros de valor
        for col_name, values in value_filters_state.items():
            # Guardar el nombre de la columna del filtro en minúsculas para consistencia
            parser.set(section_name, f'filter_{col_name.lower()}', ','.join(map(str,values)))

        # Guardar columnas seleccionadas (también en minúsculas)
        if selected_columns_state:
            parser.set(section_name, 'selected_columns', ','.join(s.lower() for s in selected_columns_state))
        else: # Si no hay columnas seleccionadas y la opción existe, quitarla
            if parser.has_option(section_name, 'selected_columns'):
                 parser.remove_option(section_name, 'selected_columns')

        # Guardar estado de "Aperturar SKU Hijo"
        parser.set(section_name, 'extend_sku_search', str(extend_sku_search_state))

        with open(CONFIG_FILE_PATH, 'w', encoding='utf-8') as configfile:
            parser.write(configfile)
        
        # Actualizar el estado en memoria (config_data["saved_states"])
        new_saved_state_in_memory = {}
        for key_opt in parser.options(section_name): # Leer con case original
            new_saved_state_in_memory[key_opt] = parser.get(section_name, key_opt, fallback='')
        config_data["saved_states"][key_upper_for_saved_state_section] = new_saved_state_in_memory

        logging.info(f"Estado de filtros guardado para '{blob_display_name}' (sección: '{section_name}') en '{CONFIG_FILE_PATH}'.")
        return True
    except Exception as e:
        logging.error(f"Error al guardar estado de filtros para '{blob_display_name}': {e}", exc_info=True)
        return False

def save_column_preference(blob_display_name: str, selected_columns: List[str]) -> bool:
    """
    Guarda las columnas seleccionadas por el usuario en SavedState del config.ini.

    Args:
        blob_display_name: Nombre de la fuente de datos
        selected_columns: Lista de columnas seleccionadas por el usuario

    Returns:
        True si se guardó exitosamente, False en caso contrario
    """
    global config_data

    parser = config_data.get("config_parser")
    if not parser:
        logging.error("ConfigParser no está cargado. No se puede guardar preferencia de columnas.")
        return False

    # La clave para la sección SavedState debe ser la versión .upper() del display_name
    key_upper_for_saved_state_section = blob_display_name.upper()
    section_name = f"SavedState_{key_upper_for_saved_state_section}"

    try:
        # Crear sección si no existe
        if not parser.has_section(section_name):
            parser.add_section(section_name)

        # Guardar columnas seleccionadas como string separado por comas (en minúsculas para consistencia)
        columns_str = ','.join(col.lower() for col in selected_columns)
        parser.set(section_name, 'last_selected_columns', columns_str)

        # Guardar en archivo
        with open(CONFIG_FILE_PATH, 'w', encoding='utf-8') as configfile:
            parser.write(configfile)

        # Actualizar el estado en memoria
        if key_upper_for_saved_state_section not in config_data["saved_states"]:
            config_data["saved_states"][key_upper_for_saved_state_section] = {}
        config_data["saved_states"][key_upper_for_saved_state_section]['last_selected_columns'] = columns_str

        logging.info(f"Preferencia de columnas guardada para '{blob_display_name}': {len(selected_columns)} columnas")
        return True

    except Exception as e:
        logging.error(f"Error guardando preferencia de columnas para '{blob_display_name}': {e}", exc_info=True)
        return False

def load_filter_state_from_config(blob_display_name: str) -> Optional[Dict[str, Any]]:
    # ... código sin cambios ...
    # blob_display_name es el original_display_name (case preservado)
    key_upper_to_load_saved_state = blob_display_name.upper() # Las claves en config_data["saved_states"] son .upper()

    logging.info(f"Cargando estado para '{blob_display_name}', usando clave UPPER '{key_upper_to_load_saved_state}' para buscar en memoria.")
    # saved_state_raw contendrá las claves con el case original del archivo config.ini
    saved_state_raw = config_data["saved_states"].get(key_upper_to_load_saved_state)

    if not saved_state_raw: # Si no hay nada en memoria (es un diccionario vacío o None)
        logging.info(f"No se encontró estado guardado en memoria para '{key_upper_to_load_saved_state}'.")
        return {"message": f"No se encontró estado guardado para {blob_display_name}.",
                "value_filters": {}, "selected_columns": [], "extend_sku_search": False}

    processed_state = {
        "value_filters": {}, # Las claves de columna aquí deben ser minúsculas
        "selected_columns": [], # Valores en minúsculas
        "extend_sku_search": False,
        "message": f"Estado cargado para {blob_display_name}."
    }
    has_any_saved_data = False # Para verificar si realmente se cargó algo
    for key_from_config, value_str in saved_state_raw.items():
        has_any_saved_data = True
        # Las claves de filtros de valor se guardaron como 'filter_nombrecolumna' (minúscula)
        if key_from_config.startswith('filter_'):
            col_name = key_from_config.split('filter_', 1)[1] # Ya debería ser minúscula
            processed_state["value_filters"][col_name] = [v.strip() for v in value_str.split(',') if v.strip() or v == ''] # Permitir strings vacíos
        # Las columnas seleccionadas se guardaron como 'selected_columns' y sus valores son una lista en minúsculas
        elif key_from_config == 'selected_columns':
            processed_state["selected_columns"] = [v.strip() for v in value_str.split(',') if v.strip()]
        elif key_from_config == 'extend_sku_search':
            processed_state["extend_sku_search"] = value_str.lower() == 'true'
    
    if not has_any_saved_data: # Si la sección SavedState existía pero estaba vacía
         processed_state["message"] = f"No hay configuraciones específicas guardadas para {blob_display_name} (sección vacía)."
           
    logging.info(f"Estado de filtros cargado desde config para '{key_upper_to_load_saved_state}': {processed_state}")
    return processed_state


# --- Funciones asíncronas seguras para acceso desde FastAPI ---
async def load_blob_data_safe(param_from_frontend_url: str) -> Dict[str, Any]:
    """Versión asíncrona de load_blob_data - lock eliminado, cache persistente es thread-safe."""
    # Crear tarea para ejecutar la carga en background con mejor control
    load_task = asyncio.create_task(
        _load_blob_data_async(param_from_frontend_url)
    )
    
    # Esperamos el resultado de la tarea
    return await load_task

async def _load_blob_data_async(param_from_frontend_url: str) -> Dict[str, Any]:
    """Versión completamente asíncrona de load_blob_data."""
    loop = asyncio.get_running_loop()
    
    # Ejecutar operaciones I/O bound en pool dedicado para mejor rendimiento
    return await loop.run_in_executor(io_executor, load_blob_data, param_from_frontend_url)


async def refresh_blob_data_safe(param_from_frontend_url: str) -> Dict[str, Any]:
    """Versión asíncrona de refresh_blob_data - lock eliminado, cache persistente es thread-safe."""
    # Crear tarea para ejecutar el refresh en background con mejor control
    refresh_task = asyncio.create_task(
        _refresh_blob_data_async(param_from_frontend_url)
    )
    
    # Esperamos el resultado de la tarea
    return await refresh_task

async def _refresh_blob_data_async(param_from_frontend_url: str) -> Dict[str, Any]:
    """Versión completamente asíncrona de refresh_blob_data."""
    loop = asyncio.get_running_loop()
    
    # Ejecutar operaciones I/O bound en pool dedicado para mejor rendimiento
    return await loop.run_in_executor(io_executor, refresh_blob_data, param_from_frontend_url)


async def apply_all_filters_safe(
    value_filters: Dict[str, List[str]],
    use_sku_hijo_file: bool,
    extend_sku_hijo: bool,
    sku_hijo_manual_list: Optional[List[str]],
    use_sku_padre_file: bool,
    sku_padre_manual_list: Optional[List[str]],
    use_ticket_file: bool,
    ticket_manual_list: Optional[List[str]],
    lineamiento_manual_list: Optional[List[str]],
    selected_display_columns: Optional[List[str]] = None,
    enable_priority_coloring: bool = False,
    page: int = 1,
    page_size: int = 100,
) -> Dict[str, Any]:
    """Versión asíncrona de apply_all_filters - lock eliminado."""
    loop = asyncio.get_running_loop()
    # Usar pool de I/O para operaciones de filtrado (accede a variables globales)
    return await loop.run_in_executor(
            io_executor,
            apply_all_filters,
            value_filters,
            use_sku_hijo_file,
            extend_sku_hijo,
            sku_hijo_manual_list,
            use_sku_padre_file,
            sku_padre_manual_list,
            use_ticket_file,
            ticket_manual_list,
            lineamiento_manual_list,
            selected_display_columns,
            enable_priority_coloring,
            page,
            page_size,
        )


async def get_excel_export_safe(
    value_filters: Dict[str, List[str]],
    use_sku_hijo_file: bool,
    extend_sku_hijo: bool,
    sku_hijo_manual_list: Optional[List[str]],
    use_sku_padre_file: bool,
    sku_padre_manual_list: Optional[List[str]],
    use_ticket_file: bool,
    ticket_manual_list: Optional[List[str]],
    lineamiento_manual_list: Optional[List[str]],
    selected_display_columns: Optional[List[str]] = None,
    enable_priority_coloring: bool = False,
) -> str:
    """Versión asíncrona de get_excel_export - lock eliminado."""
    loop = asyncio.get_running_loop()
    # Usar pool de I/O para generación de Excel (accede a variables globales)
    return await loop.run_in_executor(
            io_executor,
            get_excel_export,
            value_filters,
            use_sku_hijo_file,
            extend_sku_hijo,
            sku_hijo_manual_list,
            use_sku_padre_file,
            sku_padre_manual_list,
            use_ticket_file,
            ticket_manual_list,
            lineamiento_manual_list,
            selected_display_columns,
            enable_priority_coloring,
        )


async def get_csv_export_safe(
    value_filters: Dict[str, List[str]],
    use_sku_hijo_file: bool,
    extend_sku_hijo: bool,
    sku_hijo_manual_list: Optional[List[str]],
    use_sku_padre_file: bool,
    sku_padre_manual_list: Optional[List[str]],
    use_ticket_file: bool,
    ticket_manual_list: Optional[List[str]],
    lineamiento_manual_list: Optional[List[str]],
    selected_display_columns: Optional[List[str]] = None,
    enable_priority_coloring: bool = False,
) -> Callable[[], Iterator[bytes]]:
    """Versión asíncrona de get_csv_export - lock eliminado."""
    loop = asyncio.get_running_loop()
    # Usar pool de I/O para generación de CSV (accede a variables globales)
    return await loop.run_in_executor(
            io_executor,
            get_csv_export,
            value_filters,
            use_sku_hijo_file,
            extend_sku_hijo,
            sku_hijo_manual_list,
            use_sku_padre_file,
            sku_padre_manual_list,
            use_ticket_file,
            ticket_manual_list,
            lineamiento_manual_list,
            selected_display_columns,
            enable_priority_coloring,
        )
        

# --- INSTANCIA ÚNICA DE AUTENTICACIÓN ---
# Se crea una única instancia para que la sesión (y el token)
# se mantenga durante toda la ejecución de la aplicación.
sharepoint_authenticator = None

def get_sharepoint_authenticator():
    """Inicializa el autenticador de SharePoint de forma perezosa y lo devuelve."""
    global sharepoint_authenticator
    if sharepoint_authenticator is None:
        sharepoint_authenticator = sharepoint_auth.SharePointAuth()
    return sharepoint_authenticator

def _download_sharepoint_file_chunked(filename: str, chunk_size: int = 8192) -> bytes:
    """
    Descarga un archivo de SharePoint en chunks para optimizar memoria.
    """
    import base64
    
    auth = get_sharepoint_authenticator()
    
    try:
        access_token = auth.get_token()
    except Exception as e:
        raise ConnectionError(f"Fallo en la autenticación con SharePoint: {e}")

    # Codificar la URL de SharePoint
    if '#' in filename:
        filename = filename.split('#')[0].strip()
    
    sharepoint_url_bytes = filename.encode('utf-8')
    encoded_url = base64.urlsafe_b64encode(sharepoint_url_bytes).decode('utf-8').rstrip('=')
    
    graph_url = f"https://graph.microsoft.com/v1.0/shares/u!{encoded_url}/driveItem/content"
    
    headers = {'Authorization': 'Bearer ' + access_token}
    
    logging.info(f"Descargando archivo desde SharePoint con chunks de {chunk_size} bytes...")
    
    try:
        response = requests.get(graph_url, headers=headers, stream=True, timeout=300, verify=False)
        response.raise_for_status()
        
        # Obtener tamaño del archivo si está disponible
        content_length = response.headers.get('Content-Length')
        total_size = int(content_length) if content_length else None
        
        downloaded = 0
        content_chunks = []
        
        for chunk in response.iter_content(chunk_size=chunk_size):
            if chunk:
                content_chunks.append(chunk)
                downloaded += len(chunk)
                
                if total_size:
                    progress = (downloaded / total_size) * 100
                    if downloaded % (chunk_size * 100) == 0:  # Log cada ~800KB
                        logging.info(f"Descarga SharePoint: {progress:.1f}% ({downloaded:,}/{total_size:,} bytes)")
        
        logging.info(f"Descarga SharePoint completada: {downloaded:,} bytes")
        return b''.join(content_chunks)
        
    except requests.exceptions.RequestException as e:
        logging.error(f"Error de red al acceder a SharePoint: {e}")
        raise ConnectionError(f"Error de red al conectar con SharePoint: {e}")

def _read_file_from_sharepoint(filename: str, blob_attrs: dict, usecols: Optional[List[str]] = None) -> pd.DataFrame:
    """
    Lee un archivo desde SharePoint usando la API de Microsoft Graph con descarga optimizada.
    'filename' se espera que sea la URL completa al archivo de SharePoint.

    Args:
        filename: URL completa del archivo en SharePoint
        blob_attrs: Atributos de configuración del blob
        usecols: Lista opcional de columnas a cargar (solo para CSV, reduce uso de RAM)
    """
    # Descarga optimizada con chunks
    file_content_bytes = _download_sharepoint_file_chunked(filename)
    file_content = io.BytesIO(file_content_bytes)

    # Informar tamaño del archivo descargado
    file_size_mb = len(file_content_bytes) / (1024 * 1024)
    logging.info(f"Archivo descargado: {file_size_mb:.1f} MB. Iniciando procesamiento...")

    # Leer el archivo según su extensión
    # Obtener parámetros de configuración para esta base
    sheet_name = blob_attrs.get('sheet_name', None)  # None = primera hoja (comportamiento actual)
    values_only = blob_attrs.get('values_only', False)
    
    if filename.lower().endswith('.xlsx'):
        if sheet_name and values_only:
            # Nueva funcionalidad: hoja específica + solo valores (sin fórmulas)
            # Usar openpyxl directamente para leer solo valores
            import openpyxl
            file_content.seek(0)  # Resetear posición del archivo
            workbook = openpyxl.load_workbook(file_content, data_only=True)
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                data = []
                headers = []
                for row_num, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                    if row_num == 1:
                        # Manejar headers correctamente, eliminando celdas vacías al final
                        raw_headers = list(row)
                        # Encontrar el último índice con contenido
                        last_content_idx = -1
                        for i, cell in enumerate(raw_headers):
                            if cell is not None and str(cell).strip():
                                last_content_idx = i
                        
                        # Solo usar headers hasta el último con contenido
                        if last_content_idx >= 0:
                            headers = []
                            for i in range(last_content_idx + 1):
                                cell = raw_headers[i]
                                if cell is not None and str(cell).strip():
                                    headers.append(str(cell).strip())
                                else:
                                    headers.append(f"Column_{i}")
                        else:
                            headers = ["Column_0"]  # Fallback si no hay headers
                    else:
                        # Solo tomar los datos correspondientes al número de headers
                        row_data = list(row)[:len(headers)] if headers else list(row)
                        data.append(row_data)
                df = pd.DataFrame(data, columns=headers)
                workbook.close()
            else:
                workbook.close()
                raise ValueError(f"Hoja '{sheet_name}' no encontrada en el archivo Excel")
        elif sheet_name:
            # Solo hoja específica, mantener fórmulas
            df = pd.read_excel(file_content, engine='openpyxl', sheet_name=sheet_name)
        else:
            # Comportamiento original (NO CAMBIA NADA)
            df = pd.read_excel(file_content, engine='openpyxl')
    elif filename.lower().endswith('.csv'):
        # Mensaje específico para archivos CSV grandes
        if file_size_mb > 100:
            estimated_time = int(file_size_mb / 40)  # Estimación: ~40MB/min de parsing
            logging.info(f"⏳ Procesando archivo CSV grande ({file_size_mb:.0f} MB)...")
            logging.info(f"⏱️  Tiempo estimado: {estimated_time} minuto(s). Por favor espere, no cierre esta ventana.")

        df = csv_utils.read_csv_from_bytes(file_content.getvalue(), os.path.basename(filename), usecols=usecols)

        # Confirmar que el parsing terminó
        if file_size_mb > 100:
            logging.info(f"✅ Parsing CSV completado exitosamente.")
    else:
        # Intentar leer como Excel por defecto
        if sheet_name and values_only:
            # Usar openpyxl directamente para leer solo valores
            import openpyxl
            file_content.seek(0)  # Resetear posición del archivo
            workbook = openpyxl.load_workbook(file_content, data_only=True)
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                data = []
                headers = []
                for row_num, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                    if row_num == 1:
                        # Manejar headers correctamente, eliminando celdas vacías al final
                        raw_headers = list(row)
                        # Encontrar el último índice con contenido
                        last_content_idx = -1
                        for i, cell in enumerate(raw_headers):
                            if cell is not None and str(cell).strip():
                                last_content_idx = i
                        
                        # Solo usar headers hasta el último con contenido
                        if last_content_idx >= 0:
                            headers = []
                            for i in range(last_content_idx + 1):
                                cell = raw_headers[i]
                                if cell is not None and str(cell).strip():
                                    headers.append(str(cell).strip())
                                else:
                                    headers.append(f"Column_{i}")
                        else:
                            headers = ["Column_0"]  # Fallback si no hay headers
                    else:
                        # Solo tomar los datos correspondientes al número de headers
                        row_data = list(row)[:len(headers)] if headers else list(row)
                        data.append(row_data)
                df = pd.DataFrame(data, columns=headers)
                workbook.close()
            else:
                workbook.close()
                raise ValueError(f"Hoja '{sheet_name}' no encontrada en el archivo Excel")
        elif sheet_name:
            df = pd.read_excel(file_content, engine='openpyxl', sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_content, engine='openpyxl')
    
    return df


if not load_app_config():
    logging.critical("FALLO CRÍTICO AL CARGAR CONFIGURACIÓN INICIAL. LA APLICACIÓN PODRÍA NO FUNCIONAR CORRECTAMENTE.")

# --- VARIABLES GLOBALES PARA CANCELACIÓN DE EXPORTACIÓN ---
export_cancellation_requested = False
current_export_task = None

# --- FUNCIONES PARA CANCELACIÓN DE EXPORTACIÓN ---
def request_export_cancellation():
    """Solicita la cancelación de la exportación en curso."""
    global export_cancellation_requested
    export_cancellation_requested = True
    logging.info("Cancelación de exportación solicitada.")

def reset_export_cancellation():
    """Resetea el estado de cancelación de exportación."""
    global export_cancellation_requested
    export_cancellation_requested = False

def check_export_cancellation():
    """Verifica si se ha solicitado la cancelación de la exportación."""
    global export_cancellation_requested
    if export_cancellation_requested:
        logging.info("Exportación cancelada por el usuario.")
        raise InterruptedError("Exportación cancelada por el usuario.")
    return False

# --- FUNCIONES PARA LA FUNCIONALIDAD DE DECLARACIONES ---
def get_team_members() -> List[Dict[str, str]]:
    """
    Obtiene la lista de integrantes del equipo desde la configuración.
    Retorna una lista de diccionarios con id, display_name y excel_url.
    """
    parser = config_data.get("config_parser")
    if not parser or not parser.has_section('TeamMembers'):
        logging.warning("Sección [TeamMembers] no encontrada en config.ini")
        return []
    
    team_members = []
    for member_id, excel_url in parser.items('TeamMembers'):
        # Convertir el ID del miembro a un nombre más legible
        display_name = member_id.replace('_', ' ').title()
        team_members.append({
            'member_id': member_id,
            'display_name': display_name,
            'excel_url': excel_url
        })
    
    return team_members

def load_declaration_to_sharepoint(team_member: str, declaration_text: str) -> Dict[str, Any]:
    """
    Carga una declaración al archivo Excel de SharePoint correspondiente al integrante del equipo.
    
    Args:
        team_member: ID del integrante del equipo
        declaration_text: Texto de la declaración a cargar
    
    Returns:
        Dict con el resultado de la operación
    """
    try:
        # Obtener la configuración del integrante
        parser = config_data.get("config_parser")
        if not parser or not parser.has_section('TeamMembers'):
            raise ValueError("Sección [TeamMembers] no encontrada en config.ini")
        
        if not parser.has_option('TeamMembers', team_member):
            raise ValueError(f"Integrante '{team_member}' no encontrado en la configuración")
        
        excel_url = parser.get('TeamMembers', team_member)
        logging.info(f"URL configurada para {team_member}: {excel_url}")
        
        # Autenticar con SharePoint
        auth = get_sharepoint_authenticator()
        access_token = auth.get_token()
        
        # Leer el archivo Excel actual desde SharePoint
        logging.info(f"Leyendo archivo SharePoint para {team_member}: {excel_url}")
        df = _read_file_from_sharepoint(excel_url, {})
        logging.info(f"Archivo leído. Filas: {len(df)}, Columnas: {list(df.columns) if not df.empty else 'vacío'}")
        
        # Preparar la nueva declaración - dividir por líneas
        declaration_lines = [line.strip() for line in declaration_text.split('\n') if line.strip()]
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        current_user = team_member.replace('_', ' ').title()
        
        logging.info(f"Procesando declaración con {len(declaration_lines)} líneas")
        logging.info(f"Primeras 3 líneas: {declaration_lines[:3]}")
        
        # Crear una fila por cada línea de declaración
        declaration_data = {
            'EAN_HIJO': declaration_lines,
            'Fecha': [current_time] * len(declaration_lines),
            'Usuario': [current_user] * len(declaration_lines)
        }
        
        # Crear DataFrame con la nueva declaración
        new_declaration_df = pd.DataFrame(declaration_data)
        logging.info(f"Nueva declaración preparada con {len(new_declaration_df)} filas")
        
        # Si el archivo está vacío o no tiene la estructura esperada, crear uno nuevo
        if df.empty or 'EAN_HIJO' not in df.columns:
            logging.info("Archivo vacío o sin estructura esperada. Creando nuevo archivo.")
            df = new_declaration_df
        else:
            logging.info("Añadiendo declaración al archivo existente.")
            # Añadir la nueva declaración al final
            df = pd.concat([df, new_declaration_df], ignore_index=True)
        
        logging.info(f"DataFrame final preparado. Filas: {len(df)}, Columnas: {list(df.columns)}")
        
        # Guardar el archivo actualizado en SharePoint
        # Primero guardar temporalmente - usar directorio temporal del sistema
        import tempfile
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, f"declaration_{team_member}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        logging.info(f"Guardando archivo temporal: {temp_file}")
        df.to_excel(temp_file, index=False)
        
        # Verificar que el archivo se guardó correctamente
        if not os.path.exists(temp_file):
            raise FileNotFoundError(f"No se pudo crear el archivo temporal: {temp_file}")
        
        file_size = os.path.getsize(temp_file)
        logging.info(f"Archivo temporal creado. Tamaño: {file_size} bytes")
        
        # Subir a SharePoint usando la API de Graph
        with open(temp_file, 'rb') as file_content:
            # Codificar la URL de SharePoint
            import base64
            sharepoint_url_bytes = excel_url.encode('utf-8')
            encoded_url = base64.urlsafe_b64encode(sharepoint_url_bytes).decode('utf-8').rstrip('=')
            
            # URL para subir el contenido
            upload_url = f"https://graph.microsoft.com/v1.0/shares/u!{encoded_url}/driveItem/content"
            logging.info(f"Subiendo archivo a SharePoint: {upload_url}")
            
            headers = {
                'Authorization': 'Bearer ' + access_token,
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            file_data = file_content.read()
            logging.info(f"Datos del archivo leídos. Tamaño: {len(file_data)} bytes")
            
            response = requests.put(upload_url, headers=headers, data=file_data, verify=False)
            logging.info(f"Respuesta de SharePoint: {response.status_code} - {response.text}")
            response.raise_for_status()
        
        # Limpiar archivo temporal
        os.remove(temp_file)
        logging.info("Archivo temporal eliminado")
        
        logging.info(f"Declaración cargada exitosamente para {team_member}")
        return {
            'success': True,
            'message': f'Declaración cargada exitosamente para {team_member.replace("_", " ").title()}',
            'team_member': team_member,
            'declaration_count': len(df)
        }
        
    except Exception as e:
        logging.error(f"Error al cargar declaración para {team_member}: {e}", exc_info=True)
        return {
            'success': False,
            'message': f'Error al cargar declaración: {str(e)}',
            'team_member': team_member
        }

# --- FUNCIONES PARA LA FUNCIONALIDAD DE RECHAZOS ---
def get_team_members_reject() -> List[Dict[str, str]]:
    """
    Obtiene la lista de integrantes del equipo para rechazos desde la configuración.
    Retorna una lista de diccionarios con id, display_name y excel_url.
    """
    parser = config_data.get("config_parser")
    if not parser or not parser.has_section('TeamMembersReject'):
        logging.warning("Sección [TeamMembersReject] no encontrada en config.ini")
        return []
    
    team_members = []
    for member_id, excel_url in parser.items('TeamMembersReject'):
        # Convertir el ID del miembro a un nombre más legible
        display_name = member_id.replace('_', ' ').title()
        team_members.append({
            'member_id': member_id,
            'display_name': display_name,
            'excel_url': excel_url
        })
    
    return team_members

def load_rejection_to_sharepoint(team_member: str, rejection_text: str, rejection_obs: str) -> Dict[str, Any]:
    """
    Carga un rechazo al archivo Excel de SharePoint correspondiente al integrante del equipo.
    
    Args:
        team_member: ID del integrante del equipo
        rejection_text: Texto del rechazo a cargar
        rejection_obs: Observación del rechazo a cargar
    
    Returns:
        Dict con el resultado de la operación
    """
    try:
        # Obtener la configuración del integrante
        parser = config_data.get("config_parser")
        if not parser or not parser.has_section('TeamMembersReject'):
            raise ValueError("Sección [TeamMembersReject] no encontrada en config.ini")
        
        if not parser.has_option('TeamMembersReject', team_member):
            raise ValueError(f"Integrante '{team_member}' no encontrado en la configuración de rechazos")
        
        excel_url = parser.get('TeamMembersReject', team_member)
        
        # Autenticar con SharePoint
        auth = get_sharepoint_authenticator()
        access_token = auth.get_token()
        
        # Leer el archivo Excel actual desde SharePoint
        df = _read_file_from_sharepoint(excel_url, {})
        logging.info(f"Archivo de rechazo leído. Filas: {len(df)}, Columnas: {list(df.columns) if not df.empty else 'vacío'}")
        
        # Preparar el nuevo rechazo - dividir por líneas
        rejection_lines = [line.strip() for line in rejection_text.split('\n') if line.strip()]
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        current_user = team_member.replace('_', ' ').title()
        
        logging.info(f"Procesando rechazo con {len(rejection_lines)} líneas")
        logging.info(f"Primeras 3 líneas: {rejection_lines[:3]}")
        
        # Crear una fila por cada línea de rechazo
        rejection_data = {
            'EAN_HIJO': rejection_lines,
            'OBS_SIN_FOTO': [rejection_obs] * len(rejection_lines),
            'Fecha': [current_time] * len(rejection_lines),
            'Usuario': [current_user] * len(rejection_lines)
        }
        
        # Crear DataFrame con el nuevo rechazo
        new_rejection_df = pd.DataFrame(rejection_data)
        logging.info(f"Nuevo rechazo preparado con {len(new_rejection_df)} filas")
        
        # Si el archivo está vacío o no tiene la estructura esperada, crear uno nuevo
        if df.empty or 'EAN_HIJO' not in df.columns:
            logging.info("Archivo vacío o sin estructura esperada. Creando nuevo archivo.")
            df = new_rejection_df
        else:
            logging.info("Añadiendo rechazo al archivo existente.")
            # Asegurar que la columna OBS_SIN_FOTO existe en el DataFrame existente
            if 'OBS_SIN_FOTO' not in df.columns:
                df['OBS_SIN_FOTO'] = ''
            # Añadir el nuevo rechazo al final
            df = pd.concat([df, new_rejection_df], ignore_index=True)
        
        # Guardar el archivo actualizado en SharePoint
        # Primero guardar temporalmente
        temp_file = f"/tmp/rejection_{team_member}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        df.to_excel(temp_file, index=False)
        
        # Subir a SharePoint usando la API de Graph
        with open(temp_file, 'rb') as file_content:
            # Codificar la URL de SharePoint
            import base64
            sharepoint_url_bytes = excel_url.encode('utf-8')
            encoded_url = base64.urlsafe_b64encode(sharepoint_url_bytes).decode('utf-8').rstrip('=')
            
            # URL para subir el contenido
            upload_url = f"https://graph.microsoft.com/v1.0/shares/u!{encoded_url}/driveItem/content"
            
            headers = {
                'Authorization': 'Bearer ' + access_token,
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            response = requests.put(upload_url, headers=headers, data=file_content.read(), verify=False)
            response.raise_for_status()
        
        # Limpiar archivo temporal
        os.remove(temp_file)
        
        logging.info(f"Rechazo cargado exitosamente para {team_member}")
        return {
            'success': True,
            'message': f'Rechazo cargado exitosamente para {team_member.replace("_", " ").title()}',
            'team_member': team_member,
            'rejection_count': len(df)
        }
        
    except Exception as e:
        logging.error(f"Error al cargar rechazo para {team_member}: {e}", exc_info=True)
        return {
            'success': False,
            'message': f'Error al cargar rechazo: {str(e)}',
            'team_member': team_member
        }

# --- NUEVA FUNCIÓN UNIFICADA PARA ENCONTRAR COLUMNA ---

# --- FUNCIONES DE ENRIQUECIMIENTO DE DATOS ---
def _load_enrichment_data(enrichment_source: str, source_type: str, blob_attrs: dict) -> pd.DataFrame:
    """
    Carga los datos de enriquecimiento desde la fuente especificada.
    """
    logging.info(f"Cargando datos de enriquecimiento desde: '{enrichment_source}'")
    
    try:
        if source_type == 'azure':
            connection_string = config_data.get("connection_string")
            container_name = config_data.get("container_name")
            if not all([connection_string, container_name, enrichment_source]):
                raise ValueError("Falta la configuración de Azure para datos de enriquecimiento.")
            
            blob_service_client = BlobServiceClient.from_connection_string(connection_string)
            blob_client = blob_service_client.get_blob_client(container=container_name, blob=enrichment_source)
            downloader = blob_client.download_blob(max_concurrency=4)
            blob_content = downloader.readall()
            
            if enrichment_source.lower().endswith('.csv'):
                df_enrichment = csv_utils.read_csv_from_bytes(blob_content, enrichment_source)
            elif enrichment_source.lower().endswith(('.xlsx', '.xls')):
                df_enrichment = pd.read_excel(io.BytesIO(blob_content), engine='openpyxl')
            else:
                raise ValueError(f"Formato de archivo no soportado para enriquecimiento: {enrichment_source}")
        
        else:
            raise ValueError(f"Tipo de fuente no soportado para enriquecimiento: {source_type}")
        
        # Normalizar columnas
        df_enrichment.columns = df_enrichment.columns.str.strip().str.lower()
        logging.info(f"Datos de enriquecimiento cargados: {len(df_enrichment)} filas, {len(df_enrichment.columns)} columnas")
        
        return df_enrichment

    except Exception as e:
        logging.error(f"Error al cargar datos de enriquecimiento: {e}", exc_info=True)
        return pd.DataFrame()


def _load_enrichment_data_cached(enrichment_source: str, source_type: str, blob_attrs: dict) -> pd.DataFrame:
    """
    Carga datos de enriquecimiento con soporte de caché persistente (HITO 1.2 - OPTIMIZACIÓN).

    Esta función envuelve _load_enrichment_data con caching para evitar re-descargar
    archivos auxiliares que cambian raramente. Reduce tiempo de carga en 3-8 segundos.

    Args:
        enrichment_source: Nombre de la fuente de enriquecimiento (ej: "tabla_mejoras.csv")
        source_type: Tipo de fuente ('azure', 'sharepoint', etc.)
        blob_attrs: Atributos de configuración del blob

    Returns:
        DataFrame con datos de enriquecimiento
    """
    # Usar nombre específico para caché de enrichment
    cache_name = f"enrichment_{enrichment_source.replace('/', '_').replace('.', '_')}"

    # 1. Intentar cargar desde caché
    if persistent_cache.has_cached_data(cache_name):
        logging.info(f"⚡ [OPTIMIZACIÓN] Cargando enriquecimiento '{enrichment_source}' desde caché...")
        try:
            df_cached = persistent_cache.load_cached_data(cache_name)
            if df_cached is not None and not df_cached.empty:
                logging.info(f"✅ Enriquecimiento '{enrichment_source}' cargado desde caché ({len(df_cached)} filas, {len(df_cached.columns)} columnas)")
                return df_cached
        except Exception as e:
            logging.warning(f"⚠️ Error cargando enriquecimiento desde caché: {e}")
            persistent_cache.clear_cache(cache_name)

    # 2. Si no hay caché válido, descargar desde fuente original
    logging.info(f"Descargando enriquecimiento '{enrichment_source}' desde fuente original...")
    df_enrichment = _load_enrichment_data(enrichment_source, source_type, blob_attrs)

    # 3. Guardar en caché para futuras cargas (solo si hay datos)
    if not df_enrichment.empty:
        try:
            # Construir source_url para metadata
            source_url = enrichment_source
            if source_type == 'azure' and blob_attrs:
                container = blob_attrs.get('container_name', 'unknown')
                source_url = f"azure://{container}/{enrichment_source}"

            persistent_cache.save_cached_data(cache_name, df_enrichment, source_url)
            logging.info(f"✅ Enriquecimiento '{enrichment_source}' guardado en caché ({len(df_enrichment)} filas)")
        except Exception as e:
            logging.warning(f"⚠️ No se pudo guardar enriquecimiento en caché: {e}")

    return df_enrichment


import base64
import os
import zipfile
from pathlib import Path
from urllib.parse import quote

def _get_site_id_from_url(sharepoint_url: str, access_token: str) -> str:
    """
    Obtiene el site ID de SharePoint desde una URL completa.
    """
    import re
    
    # Extraer hostname y site path de la URL
    match = re.match(r'https://([^/]+)/sites/([^/]+)', sharepoint_url)
    if not match:
        raise ValueError(f"URL de SharePoint inválida: {sharepoint_url}")
    
    hostname = match.group(1)
    site_name = match.group(2)
    
    # Obtener site ID usando Microsoft Graph API
    graph_url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:/sites/{site_name}:"
    headers = {'Authorization': f'Bearer {access_token}'}
    
    logging.info(f"Consultando site ID con URL: {graph_url}")
    
    response = requests.get(graph_url, headers=headers, verify=False, timeout=30)
    response.raise_for_status()
    
    site_data = response.json()
    return site_data['id']

def _get_drive_id_from_site(site_id: str, access_token: str, drive_name: str = None) -> str:
    """
    Obtiene el drive ID (biblioteca de documentos) de un sitio de SharePoint.
    Si no se especifica drive_name, obtiene el drive por defecto.
    """
    graph_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    headers = {'Authorization': f'Bearer {access_token}'}
    
    logging.info(f"Consultando bibliotecas de documentos con URL: {graph_url}")
    
    response = requests.get(graph_url, headers=headers, verify=False, timeout=30)
    response.raise_for_status()
    
    drives_data = response.json()
    drives = drives_data.get('value', [])
    
    logging.info(f"Respuesta de drives: {drives_data}")
    logging.info(f"Número de drives encontrados: {len(drives)}")
    
    if drives:
        for i, drive in enumerate(drives):
            logging.info(f"Drive {i}: ID={drive.get('id')}, Name='{drive.get('name')}', Type={drive.get('driveType')}")
    
    if not drives:
        logging.error(f"No se encontraron bibliotecas de documentos. Respuesta completa: {drives_data}")
        raise ValueError("No se encontraron bibliotecas de documentos en el sitio")
    
    if drive_name:
        # Buscar drive específico por nombre
        logging.info(f"Buscando biblioteca específica: '{drive_name}'")
        for drive in drives:
            drive_name_found = drive.get('name', '')
            logging.info(f"Comparando '{drive_name_found.lower()}' con '{drive_name.lower()}'")
            if drive_name_found.lower() == drive_name.lower():
                logging.info(f"Biblioteca encontrada: {drive['id']}")
                return drive['id']
        
        # Si no se encuentra exacta, mostrar todas las disponibles
        available_drives = [drive.get('name', 'Sin nombre') for drive in drives]
        logging.error(f"Bibliotecas disponibles: {available_drives}")
        raise ValueError(f"No se encontró la biblioteca de documentos '{drive_name}'. Bibliotecas disponibles: {available_drives}")
    else:
        # Retornar el primer drive (por defecto)
        default_drive = drives[0]
        logging.info(f"Usando biblioteca por defecto: ID={default_drive['id']}, Name='{default_drive.get('name')}'")
        return default_drive['id']

def _list_folder_contents(drive_id: str, folder_path: str, access_token: str) -> list:
    """
    Lista el contenido de una carpeta en SharePoint.
    folder_path puede ser 'root' para la raíz o 'root:/ruta/a/carpeta' para subcarpetas.
    """
    if folder_path == 'root':
        graph_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/children"
    else:
        # Codificar la ruta para URL
        encoded_path = quote(folder_path)
        graph_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded_path}:/children"
    
    headers = {'Authorization': f'Bearer {access_token}'}
    
    response = requests.get(graph_url, headers=headers, verify=False, timeout=30)
    response.raise_for_status()
    
    data = response.json()
    return data.get('value', [])

def _search_folder_recursive(drive_id: str, folder_name: str, access_token: str, current_path: str = 'root') -> str:
    """
    Busca una carpeta por nombre de forma recursiva en SharePoint.
    Retorna la ruta completa de la carpeta si la encuentra, None si no.
    """
    try:
        contents = _list_folder_contents(drive_id, current_path, access_token)
        
        for item in contents:
            # Si es una carpeta y coincide el nombre
            if 'folder' in item and item['name'].lower() == folder_name.lower():
                if current_path == 'root':
                    return item['name']
                else:
                    return f"{current_path}/{item['name']}"
            
            # Si es una carpeta, buscar recursivamente
            elif 'folder' in item:
                subfolder_path = item['name'] if current_path == 'root' else f"{current_path}/{item['name']}"
                result = _search_folder_recursive(drive_id, folder_name, access_token, subfolder_path)
                if result:
                    return result
        
        return None
    except Exception as e:
        logging.warning(f"Error buscando en {current_path}: {e}")
        return None

def _download_file_from_sharepoint(drive_id: str, file_path: str, access_token: str, local_path: str):
    """
    Descarga un archivo específico desde SharePoint.
    """
    encoded_path = quote(file_path)
    graph_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded_path}:/content"
    headers = {'Authorization': f'Bearer {access_token}'}
    
    response = requests.get(graph_url, headers=headers, verify=False, timeout=60, stream=True)
    response.raise_for_status()
    
    # Crear directorio local si no existe
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    
    with open(local_path, 'wb') as f:
        for chunk in response.iter_content(chunk_size=8192):
            f.write(chunk)

def _download_folder_recursive(drive_id: str, folder_path: str, access_token: str, local_base_path: str, folder_name: str):
    """
    Descarga una carpeta completa de SharePoint de forma recursiva.
    """
    try:
        contents = _list_folder_contents(drive_id, folder_path, access_token)
        
        for item in contents:
            item_name = item['name']
            
            if 'folder' in item:
                # Es una carpeta - crear localmente y descargar recursivamente
                local_folder_path = os.path.join(local_base_path, folder_name, item_name)
                os.makedirs(local_folder_path, exist_ok=True)
                
                subfolder_path = f"{folder_path}/{item_name}" if folder_path != 'root' else item_name
                _download_folder_recursive(drive_id, subfolder_path, access_token, local_base_path, f"{folder_name}/{item_name}")
                
            elif 'file' in item:
                # Es un archivo - descargarlo
                file_path = f"{folder_path}/{item_name}" if folder_path != 'root' else item_name
                local_file_path = os.path.join(local_base_path, folder_name, item_name)
                _download_file_from_sharepoint(drive_id, file_path, access_token, local_file_path)
                
    except Exception as e:
        logging.error(f"Error descargando carpeta {folder_path}: {e}")
        raise

def search_and_download_folders_from_sharepoint(
    sharepoint_url: str,
    folder_names: list,
    download_path: str,
    drive_name: str = None
) -> dict:
    """
    Busca y descarga carpetas desde SharePoint.
    
    Args:
        sharepoint_url: URL base del sitio de SharePoint
        folder_names: Lista de nombres de carpetas a buscar
        download_path: Ruta local donde descargar las carpetas
        drive_name: Nombre de la biblioteca de documentos (opcional)
    
    Returns:
        dict: Resultado con carpetas encontradas y no encontradas
    """
    auth = get_sharepoint_authenticator()
    
    try:
        access_token = auth.get_token()
    except Exception as e:
        raise ConnectionError(f"Fallo en la autenticación con SharePoint: {e}")
    
    # Crear directorio de descarga si no existe
    os.makedirs(download_path, exist_ok=True)
    
    results = {
        'found': [],
        'not_found': [],
        'errors': []
    }
    
    try:
        # Obtener site ID y drive ID
        site_id = _get_site_id_from_url(sharepoint_url, access_token)
        drive_id = _get_drive_id_from_site(site_id, access_token, drive_name)
        
        logging.info(f"Buscando {len(folder_names)} carpetas en SharePoint...")
        
        for folder_name in folder_names:
            try:
                logging.info(f"Buscando carpeta: {folder_name}")
                
                # Buscar la carpeta
                folder_path = _search_folder_recursive(drive_id, folder_name, access_token)
                
                if folder_path:
                    logging.info(f"Carpeta encontrada: {folder_name} en {folder_path}")
                    
                    # Descargar la carpeta
                    _download_folder_recursive(drive_id, folder_path, access_token, download_path, folder_name)
                    
                    results['found'].append({
                        'name': folder_name,
                        'path': folder_path,
                        'downloaded_to': os.path.join(download_path, folder_name)
                    })
                    
                    logging.info(f"Carpeta {folder_name} descargada exitosamente")
                else:
                    logging.warning(f"Carpeta no encontrada: {folder_name}")
                    results['not_found'].append(folder_name)
                    
            except Exception as e:
                error_msg = f"Error procesando carpeta {folder_name}: {str(e)}"
                logging.error(error_msg)
                results['errors'].append(error_msg)
    
    except Exception as e:
        error_msg = f"Error general en SharePoint: {str(e)}"
        logging.error(error_msg)
        results['errors'].append(error_msg)
    
    return results

# Variable global para el buscador integral de carpetas
folder_search_excel_data: Optional[bytes] = None

# Configuración del sitio SharePoint por defecto
DEFAULT_SHAREPOINT_SITE = "https://ripleycorp.sharepoint.com/sites/Publicacinycontenido2"

def get_default_sharepoint_libraries() -> dict:
    """
    Obtiene las bibliotecas de documentos disponibles en el sitio SharePoint por defecto.
    """
    auth = get_sharepoint_authenticator()
    
    try:
        access_token = auth.get_token()
    except Exception as e:
        raise ConnectionError(f"Fallo en la autenticación con SharePoint: {e}")
    
    try:
        # Obtener site ID del sitio por defecto
        site_id = _get_site_id_from_url(DEFAULT_SHAREPOINT_SITE, access_token)
        
        libraries = []
        
        # Método 1: Intentar obtener drives (bibliotecas de documentos)
        graph_url_drives = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
        headers = {'Authorization': f'Bearer {access_token}'}
        
        logging.info(f"Consultando bibliotecas con URL (drives): {graph_url_drives}")
        
        try:
            response = requests.get(graph_url_drives, headers=headers, verify=False, timeout=30)
            response.raise_for_status()
            
            drives_data = response.json()
            drives = drives_data.get('value', [])
            
            logging.info(f"Respuesta de drives: {len(drives)} encontrados")
            
            for drive in drives:
                libraries.append({
                    'id': drive.get('id'),
                    'name': drive.get('name'),
                    'description': drive.get('description', ''),
                    'type': drive.get('driveType', 'documentLibrary'),
                    'source': 'drives'
                })
        except Exception as e:
            logging.warning(f"Error obteniendo drives: {e}")
        
        # Método 2: Si no hay drives, intentar con lists (incluye bibliotecas de documentos)
        if not libraries:
            graph_url_lists = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists"
            
            logging.info(f"Consultando bibliotecas con URL (lists): {graph_url_lists}")
            
            try:
                response = requests.get(graph_url_lists, headers=headers, verify=False, timeout=30)
                response.raise_for_status()
                
                lists_data = response.json()
                lists = lists_data.get('value', [])
                
                logging.info(f"Respuesta de lists: {len(lists)} encontradas")
                logging.info(f"Respuesta completa de lists: {lists_data}")
                
                # Filtrar solo las bibliotecas de documentos
                for list_item in lists:
                    list_template = list_item.get('list', {}).get('template', '')
                    list_name = list_item.get('displayName', list_item.get('name', ''))
                    
                    # Incluir bibliotecas de documentos y algunas otras útiles
                    if list_template in ['documentLibrary', 'pictureLibrary', 'assetLibrary'] or 'document' in list_name.lower():
                        libraries.append({
                            'id': list_item.get('id'),
                            'name': list_name,
                            'description': list_item.get('description', ''),
                            'type': list_template or 'documentLibrary',
                            'source': 'lists',
                            'webUrl': list_item.get('webUrl', '')
                        })
                        logging.info(f"Biblioteca encontrada: {list_name} (template: {list_template})")
            except Exception as e:
                logging.warning(f"Error obteniendo lists: {e}")
        
        # Método 3: Si aún no hay bibliotecas, mostrar todas las listas disponibles para diagnóstico
        if not libraries:
            logging.warning("No se encontraron bibliotecas de documentos. Mostrando todas las listas para diagnóstico:")
            try:
                response = requests.get(graph_url_lists, headers=headers, verify=False, timeout=30)
                if response.status_code == 200:
                    lists_data = response.json()
                    all_lists = lists_data.get('value', [])
                    
                    for list_item in all_lists:
                        list_template = list_item.get('list', {}).get('template', 'unknown')
                        list_name = list_item.get('displayName', list_item.get('name', 'Sin nombre'))
                        
                        logging.info(f"Lista disponible: '{list_name}' (template: {list_template})")
                        
                        # Agregar todas las listas para que el usuario pueda seleccionar
                        libraries.append({
                            'id': list_item.get('id'),
                            'name': f"{list_name} [{list_template}]",
                            'description': list_item.get('description', ''),
                            'type': list_template,
                            'source': 'all_lists',
                            'webUrl': list_item.get('webUrl', '')
                        })
            except Exception as e:
                logging.error(f"Error obteniendo todas las listas: {e}")
        
        return {
            'success': True,
            'site_url': DEFAULT_SHAREPOINT_SITE,
            'site_id': site_id,
            'libraries': libraries,
            'total_found': len(libraries)
        }
        
    except Exception as e:
        logging.error(f"Error obteniendo bibliotecas de SharePoint: {e}")
        return {
            'success': False,
            'error': str(e),
            'site_url': DEFAULT_SHAREPOINT_SITE
        }

def browse_sharepoint_folder(drive_id: str, folder_path: str = 'root') -> dict:
    """
    Navega por una carpeta específica en SharePoint y retorna su contenido.
    """
    auth = get_sharepoint_authenticator()
    
    try:
        access_token = auth.get_token()
    except Exception as e:
        raise ConnectionError(f"Fallo en la autenticación con SharePoint: {e}")
    
    try:
        # Primero intentar como drive tradicional
        try:
            contents = _list_folder_contents(drive_id, folder_path, access_token)
        except Exception as drive_error:
            logging.warning(f"Error navegando como drive: {drive_error}")
            # Si falla, intentar como lista de SharePoint
            contents = _list_sharepoint_list_contents(drive_id, folder_path, access_token)
        
        folders = []
        files = []
        
        for item in contents:
            item_info = {
                'name': item.get('name'),
                'id': item.get('id'),
                'size': item.get('size', 0),
                'lastModified': item.get('lastModifiedDateTime'),
                'path': f"{folder_path}/{item.get('name')}" if folder_path != 'root' else item.get('name')
            }
            
            if 'folder' in item:
                item_info['type'] = 'folder'
                item_info['childCount'] = item.get('folder', {}).get('childCount', 0)
                folders.append(item_info)
            elif 'file' in item:
                item_info['type'] = 'file'
                item_info['mimeType'] = item.get('file', {}).get('mimeType', '')
                files.append(item_info)
        
        return {
            'success': True,
            'current_path': folder_path,
            'folders': folders,
            'files': files,
            'total_items': len(contents)
        }
        
    except Exception as e:
        logging.error(f"Error navegando carpeta {folder_path}: {e}")
        return {
            'success': False,
            'error': str(e),
            'current_path': folder_path
        }

def _list_sharepoint_list_contents(list_id: str, folder_path: str, access_token: str) -> list:
    """
    Lista el contenido de una lista de SharePoint (alternativa cuando no es un drive tradicional).
    """
    # Obtener site ID
    site_id = _get_site_id_from_url(DEFAULT_SHAREPOINT_SITE, access_token)
    
    if folder_path == 'root':
        # Para listas, obtener los elementos de la raíz
        graph_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/drive/root/children"
    else:
        # Para subcarpetas en listas
        encoded_path = quote(folder_path)
        graph_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{list_id}/drive/root:/{encoded_path}:/children"
    
    headers = {'Authorization': f'Bearer {access_token}'}
    
    logging.info(f"Consultando contenido de lista: {graph_url}")
    
    response = requests.get(graph_url, headers=headers, verify=False, timeout=30)
    response.raise_for_status()
    
    data = response.json()
    return data.get('value', [])

def search_and_download_folders_from_sharepoint_drive(
    drive_id: str,
    start_folder_path: str,
    folder_names: list,
    download_path: str
) -> dict:
    """
    Busca y descarga carpetas desde SharePoint usando drive ID y carpeta de inicio.
    
    Args:
        drive_id: ID de la biblioteca de documentos de SharePoint
        start_folder_path: Ruta de la carpeta donde empezar a buscar
        folder_names: Lista de nombres de carpetas a buscar
        download_path: Ruta local donde descargar las carpetas
    
    Returns:
        dict: Resultado con carpetas encontradas y no encontradas
    """
    auth = get_sharepoint_authenticator()
    
    try:
        access_token = auth.get_token()
    except Exception as e:
        raise ConnectionError(f"Fallo en la autenticación con SharePoint: {e}")
    
    # Crear directorio de descarga si no existe
    os.makedirs(download_path, exist_ok=True)
    
    results = {
        'found': [],
        'not_found': [],
        'errors': []
    }
    
    try:
        logging.info(f"Buscando {len(folder_names)} carpetas en drive {drive_id} desde {start_folder_path}...")
        
        for folder_name in folder_names:
            try:
                logging.info(f"Buscando carpeta: {folder_name}")
                
                # Buscar la carpeta desde la ubicación especificada
                folder_path = _search_folder_recursive_from_path(drive_id, folder_name, access_token, start_folder_path)
                
                if folder_path:
                    logging.info(f"Carpeta encontrada: {folder_name} en {folder_path}")
                    
                    # Descargar la carpeta
                    _download_folder_recursive(drive_id, folder_path, access_token, download_path, folder_name)
                    
                    results['found'].append({
                        'name': folder_name,
                        'path': folder_path,
                        'downloaded_to': os.path.join(download_path, folder_name)
                    })
                    
                    logging.info(f"Carpeta {folder_name} descargada exitosamente")
                else:
                    logging.warning(f"Carpeta no encontrada: {folder_name}")
                    results['not_found'].append(folder_name)
                    
            except Exception as e:
                error_msg = f"Error procesando carpeta {folder_name}: {str(e)}"
                logging.error(error_msg)
                results['errors'].append(error_msg)
    
    except Exception as e:
        error_msg = f"Error general en búsqueda de SharePoint: {str(e)}"
        logging.error(error_msg)
        results['errors'].append(error_msg)
    
    return results

def _search_folder_recursive_from_path(drive_id: str, folder_name: str, access_token: str, start_path: str) -> str:
    """
    Busca una carpeta por nombre de forma recursiva en SharePoint desde una ruta específica.
    Retorna la ruta completa de la carpeta si la encuentra, None si no.
    """
    try:
        contents = _list_folder_contents(drive_id, start_path, access_token)
        
        for item in contents:
            # Si es una carpeta y coincide el nombre
            if 'folder' in item and item['name'].lower() == folder_name.lower():
                if start_path == 'root':
                    return item['name']
                else:
                    return f"{start_path}/{item['name']}"
            
            # Si es una carpeta, buscar recursivamente
            elif 'folder' in item:
                subfolder_path = item['name'] if start_path == 'root' else f"{start_path}/{item['name']}"
                result = _search_folder_recursive_from_path(drive_id, folder_name, access_token, subfolder_path)
                if result:
                    return result
        
        return None
    except Exception as e:
        logging.warning(f"Error buscando en {start_path}: {e}")
        return None

def search_and_download_folders_from_sharepoint_manual(
    sharepoint_url: str,
    folder_names: list,
    download_path: str
) -> dict:
    auth = get_sharepoint_authenticator()
    try:
        access_token = auth.get_token()
    except Exception as e:
        sse_progress(f"ERR: Fallo en la autenticación con SharePoint: {e}")
        raise ConnectionError(f"Fallo en la autenticación con SharePoint: {e}")
    os.makedirs(download_path, exist_ok=True)
    results = {'found': [], 'not_found': [], 'errors': []}
    try:
        sse_progress(f"NAV: Procesando URL manual: {sharepoint_url}")
        site_base_url, folder_path = _parse_sharepoint_url(sharepoint_url)
        sse_progress(f"NAV: URL base del sitio: {site_base_url}")
        sse_progress(f"NAV: Carpeta de inicio: {folder_path}")
        site_id = _get_site_id_from_url(site_base_url, access_token)
        sse_progress(f"NAV: Buscando {len(folder_names)} carpetas desde {folder_path}...")
        for folder_name in folder_names:
            try:
                sse_progress(f"NAV: Buscando carpeta: {folder_name}")
                found_folder_url = _search_folder_by_url(site_id, folder_path, folder_name, access_token)
                if found_folder_url:
                    sse_progress(f"FOUND: {folder_name}")
                    sse_progress(f"DL: Descargando {folder_name}")
                    _download_folder_from_url(found_folder_url, access_token, download_path, folder_name)
                    results['found'].append({
                        'name': folder_name,
                        'path': found_folder_url,
                        'downloaded_to': os.path.join(download_path, folder_name)
                    })
                    sse_progress(f"DL: Carpeta {folder_name} descargada exitosamente")
                else:
                    sse_progress(f"NOTFOUND: {folder_name}")
                    results['not_found'].append(folder_name)
            except Exception as e:
                error_msg = f"Error procesando carpeta {folder_name}: {str(e)}"
                sse_progress(f"ERR: {error_msg}")
                results['errors'].append(error_msg)
    except Exception as e:
        error_msg = f"Error general en búsqueda manual de SharePoint: {str(e)}"
        sse_progress(f"ERR: {error_msg}")
        results['errors'].append(error_msg)
    return results


def _search_folder_by_url(site_id: str, base_folder_path: str, folder_name: str, access_token: str) -> str:
    try:
        # Intentar acceso directo a la carpeta
        if base_folder_path == 'root':
            search_path = folder_name
        else:
            search_path = f"{base_folder_path}/{folder_name}"
        from urllib.parse import quote
        encoded_path = quote(search_path)
        graph_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{encoded_path}"
        headers = {'Authorization': f'Bearer {access_token}'}
        logging.info(f"[DEBUG] Consultando carpeta directa: {graph_url}")
        sse_progress(f"DEBUG: Consultando {search_path}")
        response = requests.get(graph_url, headers=headers, verify=False, timeout=30)
        logging.info(f"[DEBUG] Respuesta directa: {response.status_code} {response.text[:200]}")
        if response.status_code == 200:
            item_data = response.json()
            if 'folder' in item_data:
                return search_path
        # Si no se encuentra directamente, buscar recursivamente
        return _search_folder_recursive_in_path(site_id, base_folder_path, folder_name, access_token)
    except Exception as e:
        logging.warning(f"Error buscando carpeta {folder_name}: {e}")
        sse_progress(f"DEBUG: Error buscando {search_path}: {e}")
        return None

def _search_folder_recursive_in_path(site_id: str, base_path: str, folder_name: str, access_token: str) -> str:
    try:
        from urllib.parse import quote
        if base_path == 'root':
            graph_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root/children"
        else:
            encoded_path = quote(base_path)
            graph_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{encoded_path}:/children"
        headers = {'Authorization': f'Bearer {access_token}'}
        logging.info(f"[DEBUG] Listando hijos: {graph_url}")
        sse_progress(f"DEBUG: Listando hijos en {base_path}")
        response = requests.get(graph_url, headers=headers, verify=False, timeout=30)
        logging.info(f"[DEBUG] Respuesta hijos: {response.status_code} {response.text[:200]}")
        response.raise_for_status()
        data = response.json()
        items = data.get('value', [])
        for item in items:
            if 'folder' in item and item['name'].lower() == folder_name.lower():
                if base_path == 'root':
                    return item['name']
                else:
                    return f"{base_path}/{item['name']}"
            elif 'folder' in item:
                subfolder_path = item['name'] if base_path == 'root' else f"{base_path}/{item['name']}"
                result = _search_folder_recursive_in_path(site_id, subfolder_path, folder_name, access_token)
                if result:
                    return result
        return None
    except Exception as e:
        logging.warning(f"Error en búsqueda recursiva en {base_path}: {e}")
        sse_progress(f"DEBUG: Error recursivo en {base_path}: {e}")
        return None

async def _download_folder_from_url(folder_path: str, access_token: str, local_base_path: str, folder_name: str):
    """
    Descarga una carpeta completa usando su ruta en SharePoint.
    """
    site_id = _get_site_id_from_url(DEFAULT_SHAREPOINT_SITE, access_token)
    
    try:
        from urllib.parse import quote
        encoded_path = quote(folder_path)
        
        graph_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{encoded_path}:/children"
        headers = {'Authorization': f'Bearer {access_token}'}
        
        response = requests.get(graph_url, headers=headers, verify=False, timeout=30)
        response.raise_for_status()
        
        data = response.json()
        items = data.get('value', [])
        
        for item in items:
            item_name = item['name']
            
            if 'folder' in item:
                # Es una carpeta - crear localmente y descargar recursivamente
                local_folder_path = os.path.join(local_base_path, folder_name, item_name)
                os.makedirs(local_folder_path, exist_ok=True)
                
                subfolder_path = f"{folder_path}/{item_name}"
                _download_folder_from_url(subfolder_path, access_token, local_base_path, f"{folder_name}/{item_name}")
                
            elif 'file' in item:
                # Es un archivo - descargarlo
                file_download_url = item.get('@microsoft.graph.downloadUrl')
                if file_download_url:
                    local_file_path = os.path.join(local_base_path, folder_name, item_name)
                    _download_file_from_download_url(file_download_url, local_file_path)
                
    except Exception as e:
        logging.error(f"Error descargando carpeta {folder_path}: {e}")
        raise


import asyncio


# En la función search_and_download_folders_from_sharepoint_manual, agregar llamadas a sse_progress:
# Ejemplo:
# sse_progress(f"Navegando: {folder_path}")
# sse_progress(f"Buscando carpeta: {folder_name}")
# sse_progress(f"Carpeta encontrada: {folder_name}")
# sse_progress(f"Descargando: {folder_name}")
# sse_progress(f"Carpeta no encontrada: {folder_name}")
# sse_progress(f"Error: {error_msg}")


# ===== FUNCIONES DE FAVORITOS (Sistema independiente de puerto) =====

def save_favorite_to_config(state: dict, database_name: str) -> bool:
    """
    Guarda el favorito en config.ini bajo la sección [Favorite_{database_name}].
    Cada base de datos tiene su propio favorito independiente.

    Args:
        state: Dict con value_filters, selected_columns, extend_sku_search
        database_name: Nombre de la base de datos (ej: "Chile_Wop", "Peru_Staff")

    Returns:
        bool: True si se guardó exitosamente
    """
    global config_data

    parser = config_data.get("config_parser")
    if not parser:
        logging.error("ConfigParser no está cargado. No se puede guardar el favorito.")
        return False

    # Crear nombre de sección específico para esta base
    section_name = f"Favorite_{database_name}"

    try:
        # Crear sección si no existe
        if not parser.has_section(section_name):
            parser.add_section(section_name)
        else:
            # Limpiar sección existente antes de guardar nuevos valores
            for key in parser.options(section_name):
                parser.remove_option(section_name, key)

        # Guardar datos como JSON strings
        parser.set(section_name, 'value_filters', json.dumps(state.get('value_filters', {})))
        parser.set(section_name, 'selected_columns', json.dumps(state.get('selected_columns', [])))
        parser.set(section_name, 'extend_sku_search', str(state.get('extend_sku_search', False)))
        parser.set(section_name, 'database_name', database_name)
        parser.set(section_name, 'timestamp', datetime.now().isoformat())

        # Escribir a archivo
        with open(CONFIG_FILE_PATH, 'w', encoding='utf-8') as configfile:
            parser.write(configfile)

        logging.info(f"Favorito guardado exitosamente en '{CONFIG_FILE_PATH}' (sección: '{section_name}').")
        return True

    except Exception as e:
        logging.error(f"Error al guardar favorito: {e}", exc_info=True)
        return False


def load_favorite_from_config(database_name: str) -> Optional[Dict[str, Any]]:
    """
    Carga el favorito desde config.ini para la base de datos específica.
    Cada base tiene su propio favorito independiente.

    Args:
        database_name: Nombre de la base de datos (ej: "Chile_Wop", "Peru_Staff")

    Returns:
        dict | None: Favorito guardado con metadata o None si no existe
    """
    global config_data

    parser = config_data.get("config_parser")
    if not parser:
        logging.error("ConfigParser no está cargado. No se puede cargar el favorito.")
        return None

    # Buscar sección específica para esta base
    section_name = f"Favorite_{database_name}"

    try:
        if not parser.has_section(section_name):
            logging.info(f"No existe sección [{section_name}] en config.ini")
            return None

        # Leer datos
        value_filters_json = parser.get(section_name, 'value_filters', fallback='{}')
        selected_columns_json = parser.get(section_name, 'selected_columns', fallback='[]')
        extend_sku_search = parser.getboolean(section_name, 'extend_sku_search', fallback=False)
        saved_database_name = parser.get(section_name, 'database_name', fallback=database_name)
        timestamp = parser.get(section_name, 'timestamp', fallback=None)

        # Parsear JSON
        value_filters = json.loads(value_filters_json)
        selected_columns = json.loads(selected_columns_json)

        favorite = {
            'value_filters': value_filters,
            'selected_columns': selected_columns,
            'extend_sku_search': extend_sku_search,
            'database_name': saved_database_name,
            'timestamp': timestamp
        }

        logging.info(f"Favorito cargado exitosamente desde '{CONFIG_FILE_PATH}' (sección: '{section_name}').")
        return favorite

    except json.JSONDecodeError as e:
        logging.error(f"Error al decodificar JSON del favorito: {e}", exc_info=True)
        return None
    except Exception as e:
        logging.error(f"Error al cargar favorito: {e}", exc_info=True)
        return None


def delete_favorite_from_config(database_name: str) -> bool:
    """
    Elimina el favorito de config.ini para la base de datos específica.

    Args:
        database_name: Nombre de la base de datos (ej: "Chile_Wop", "Peru_Staff")

    Returns:
        bool: True si se eliminó, False si no existía
    """
    global config_data

    parser = config_data.get("config_parser")
    if not parser:
        logging.error("ConfigParser no está cargado. No se puede eliminar el favorito.")
        return False

    # Eliminar sección específica para esta base
    section_name = f"Favorite_{database_name}"

    try:
        if not parser.has_section(section_name):
            logging.info(f"No existe sección [{section_name}] para eliminar")
            return False

        # Eliminar sección
        parser.remove_section(section_name)

        # Escribir a archivo
        with open(CONFIG_FILE_PATH, 'w', encoding='utf-8') as configfile:
            parser.write(configfile)

        logging.info(f"Favorito eliminado exitosamente de '{CONFIG_FILE_PATH}' (sección: '{section_name}').")
        return True

    except Exception as e:
        logging.error(f"Error al eliminar favorito: {e}", exc_info=True)
        return False


def has_favorite_for_database(database_name: str) -> Optional[Dict[str, Any]]:
    """
    Verifica si existe un favorito para la base de datos específica.

    Args:
        database_name: Nombre de la base de datos (ej: "Chile_Wop", "Peru_Staff")

    Returns:
        dict | None: Metadata básica del favorito (database_name, timestamp) o None si no existe
    """
    global config_data

    parser = config_data.get("config_parser")
    if not parser:
        logging.error("ConfigParser no está cargado. No se puede verificar el favorito.")
        return None

    section_name = f"Favorite_{database_name}"

    try:
        if not parser.has_section(section_name):
            return None

        # Obtener solo metadata básica
        saved_database_name = parser.get(section_name, 'database_name', fallback=database_name)
        timestamp = parser.get(section_name, 'timestamp', fallback=None)

        return {
            'exists': True,
            'database_name': saved_database_name,
            'timestamp': timestamp
        }

    except Exception as e:
        logging.error(f"Error al verificar favorito: {e}", exc_info=True)
        return None
