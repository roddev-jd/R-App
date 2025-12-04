import os
import shutil
import threading
import queue
import tkinter as tk
from tkinter import filedialog, messagebox
import concurrent.futures  # ### CAMBIO v2.5: Filtrado por depto

import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.scrolled import ScrolledText

import openpyxl  # pip install openpyxl
import csv  # Módulo csv es parte de la librería estándar de Python
# Función de tema compatible multiplataforma
def setup_theme(window=None):
    import ttkbootstrap as ttk
    style = ttk.Style()
    
    # Lista de temas ordenados por compatibilidad multiplataforma
    preferred_themes = ["flatly", "litera", "cosmo", "journal", "sandstone"]
    available = list(style.theme_names())
    
    for theme in preferred_themes:
        if theme in available:
            try:
                style.theme_use(theme)
                return style
            except:
                continue
    
    # Fallback a tema por defecto
    try:
        style.theme_use("default")
    except:
        pass
    
    return style


class BuscadorCarpetasApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Buscador de carpetas")
        self.master.geometry("800x750")

        # Variables de estado
        self.ruta_principal = tk.StringVar()
        self.archivo_planilla = tk.StringVar()
        self.nombre_columna_planilla = tk.StringVar()
        self.ruta_destino = tk.StringVar()
        self.usar_mismo_destino = tk.BooleanVar(value=False)
        self.filtro_depto = tk.StringVar(value="TODOS")
        self.depto_options = []
        self.cancelar_busqueda = False
        self.progress_text = tk.StringVar(value="Progreso: 0%")

        self.queue = queue.Queue()
        self.crear_interfaz()
        self.master.after(100, self.check_queue)

    def crear_interfaz(self):
        frm_form = ttk.Frame(self.master, padding=(20, 10))
        frm_form.grid(row=0, column=0, padx=10, pady=10, sticky=EW)
        frm_form.columnconfigure(1, weight=1)

        # Ruta principal
        ttk.Label(frm_form, text="Ruta principal:").grid(row=0, column=0, sticky=W, pady=5)
        ttk.Entry(frm_form, textvariable=self.ruta_principal, width=60).grid(row=0, column=1, sticky=EW, pady=5)
        ttk.Button(frm_form, text="Examinar", command=self.seleccionar_ruta, bootstyle="secondary").grid(row=0, column=2, padx=(5, 0))

        # Archivo planilla
        ttk.Label(frm_form, text="Archivo con nombres (Excel/CSV):").grid(row=1, column=0, sticky=W, pady=5)
        ttk.Entry(frm_form, textvariable=self.archivo_planilla, width=60).grid(row=1, column=1, sticky=EW, pady=5)
        ttk.Button(frm_form, text="Examinar", command=self.seleccionar_archivo, bootstyle="secondary").grid(row=1, column=2, padx=(5, 0))

        # Nombre columna (nombres)
        ttk.Label(frm_form, text="Nombre Columna (Opcional, defecto: 1ra):").grid(row=2, column=0, sticky=W, pady=5)
        ttk.Entry(frm_form, textvariable=self.nombre_columna_planilla, width=60).grid(row=2, column=1, sticky=EW, pady=5)

        # Filtro depto
        ttk.Label(frm_form, text="Filtrar Depto:").grid(row=3, column=0, sticky=W, pady=5)
        self.combo_depto = ttk.Combobox(frm_form, textvariable=self.filtro_depto, state='readonly', width=58)
        self.combo_depto['values'] = ['TODOS']
        self.combo_depto.current(0)
        self.combo_depto.grid(row=3, column=1, sticky=EW, pady=5)

        # Usar mismo destino
        ttk.Checkbutton(
            frm_form,
            text="Usar carpeta de planilla como destino",
            variable=self.usar_mismo_destino,
            command=self.actualizar_destino_auto,
            bootstyle="primary-round-toggle"
        ).grid(row=4, column=1, columnspan=2, sticky=W, pady=(10, 5))

        # Carpeta destino
        ttk.Label(frm_form, text="Carpeta de destino:").grid(row=5, column=0, sticky=W, pady=5)
        self.entry_destino = ttk.Entry(frm_form, textvariable=self.ruta_destino, width=60)
        self.entry_destino.grid(row=5, column=1, sticky=EW, pady=5)
        ttk.Button(frm_form, text="Examinar", command=self.seleccionar_destino, bootstyle="secondary").grid(row=5, column=2)

        # Botones iniciar/cancelar
        frm_buttons = ttk.Frame(self.master, padding=(0, 10))
        frm_buttons.grid(row=1, column=0, pady=(5, 10), sticky=E, padx=20)
        self.btn_iniciar = ttk.Button(frm_buttons, text="Iniciar búsqueda y copia", command=self.iniciar_busqueda, bootstyle="success")
        self.btn_iniciar.pack(side=RIGHT, padx=(5, 0))
        self.btn_cancelar = ttk.Button(frm_buttons, text="Cancelar búsqueda", command=self.cancelar_busqueda_func, bootstyle="danger-outline", state=DISABLED)
        self.btn_cancelar.pack(side=RIGHT, padx=(10, 5))

        # Barra de progreso y log
        frm_progress = ttk.Frame(self.master, padding=(0, 5))
        frm_progress.grid(row=2, column=0, padx=20, pady=(0, 5), sticky=EW)
        self.barra_progreso = ttk.Progressbar(frm_progress, orient=HORIZONTAL, length=400, mode="determinate", bootstyle="info-striped")
        self.barra_progreso.pack(fill=X, pady=(0, 5))
        self.lbl_progress_text = ttk.Label(frm_progress, textvariable=self.progress_text)
        self.lbl_progress_text.pack(anchor=W)

        lf_log = ttk.LabelFrame(self.master, text="Recorrido de directorios", padding=10, bootstyle="info")
        lf_log.grid(row=3, column=0, padx=20, pady=(5, 10), sticky=NSEW)
        self.txt_recorrido = ScrolledText(lf_log, wrap=WORD, height=10, autohide=True, hbar=False)
        self.txt_recorrido.pack(fill=BOTH, expand=True, padx=5, pady=5)
        self.txt_recorrido.text.config(state=tk.DISABLED)
        self.master.rowconfigure(3, weight=1)
        self.master.columnconfigure(0, weight=1)

    def seleccionar_ruta(self):
        ruta = filedialog.askdirectory(title="Seleccionar carpeta principal")
        if ruta:
            self.ruta_principal.set(ruta)

    def seleccionar_archivo(self):
        archivo = filedialog.askopenfilename(
            title="Seleccionar archivo de planilla",
            filetypes=[("Archivos de Excel", "*.xlsx *.xls"), ("Archivos CSV", "*.csv"), ("Todos los archivos", "*.*")]
        )
        if archivo:
            self.archivo_planilla.set(archivo)
            self.actualizar_destino_auto()
            self._cargar_opciones_depto()

    def _cargar_opciones_depto(self):
        """Lee la columna 'depto' de la planilla y actualiza el combobox"""
        try:
            archivo = self.archivo_planilla.get()
            ext = os.path.splitext(archivo)[1].lower()
            deptos = set()
            if ext in ['.xlsx', '.xls']:
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active
                headers = [str(cell.value or '').strip().lower() for cell in sheet[1]]
                if 'depto' not in headers:
                    raise ValueError("No se encontró la columna 'depto' en la planilla.")
                idx = headers.index('depto')
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > idx and row[idx]:
                        deptos.add(str(row[idx]).strip())
            elif ext == '.csv':
                with open(archivo, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    headers = [h.strip().lower() for h in next(reader)]
                    if 'depto' not in headers:
                        raise ValueError("No se encontró la columna 'depto' en la planilla.")
                    idx = headers.index('depto')
                    for row in reader:
                        if row and len(row) > idx and row[idx]:
                            deptos.add(row[idx].strip())
            else:
                return
            self.depto_options = ['TODOS'] + sorted(deptos)
            self.combo_depto['values'] = self.depto_options
            self.combo_depto.current(0)
        except Exception as e:
            messagebox.showerror("Error de Planilla", str(e))

    def seleccionar_destino(self):
        destino = filedialog.askdirectory(title="Seleccionar carpeta de destino")
        if destino:
            self.ruta_destino.set(destino)

    def actualizar_destino_auto(self):
        if self.usar_mismo_destino.get() and self.archivo_planilla.get():
            self.ruta_destino.set(os.path.dirname(self.archivo_planilla.get()))
            self.entry_destino.config(state=DISABLED)
        else:
            self.entry_destino.config(state=NORMAL)

    def iniciar_busqueda(self):
        self.cancelar_busqueda = False
        ruta = self.ruta_principal.get().strip()
        archivo = self.archivo_planilla.get().strip()
        destino = self.ruta_destino.get().strip()
        nombre_col = self.nombre_columna_planilla.get().strip() or None
        filtro = self.filtro_depto.get()

        # Validación de entradas
        if not all([ruta, os.path.isdir(ruta), archivo, os.path.isfile(archivo), destino, os.path.isdir(destino)]):
            messagebox.showerror("Error", "Por favor, verifique que todas las rutas sean válidas.")
            return

        # Lectura planilla
        nombres, deptos = self.leer_nombres_carpetas(archivo, nombre_col)
        if not nombres:
            messagebox.showerror("Error", "No se encontraron nombres de carpeta en la planilla.")
            return

        # Filtrado por depto
        if filtro != 'TODOS':
            nombres = [n for n, d in zip(nombres, deptos) if d == filtro]
            if not nombres:
                messagebox.showinfo("Sin datos", f"No hay filas con depto = {filtro}.")
                return

        # Preparar interfaz
        self.txt_recorrido.text.config(state=tk.NORMAL)
        self.txt_recorrido.text.delete("1.0", tk.END)
        self.txt_recorrido.text.config(state=tk.DISABLED)
        self.barra_progreso["value"] = 0
        self.progress_text.set("Progreso: 0%")
        self.btn_iniciar.config(state=DISABLED)
        self.btn_cancelar.config(state=NORMAL)
        self.barra_progreso.configure(bootstyle="info-striped")

        # Iniciar hilo de trabajo
        hilo = threading.Thread(target=self.trabajo_en_hilo, args=(nombres, ruta, destino), daemon=True)
        hilo.start()

    def cancelar_busqueda_func(self):
        self.cancelar_busqueda = True

    def _copiar_una_carpeta(self, nombre_original, ruta_encontrada, destino_base):
        try:
            self.queue.put(("log", f"COPIANDO '{nombre_original}'..."))
            base = os.path.basename(ruta_encontrada)
            dest = os.path.join(destino_base, base)
            count = 2
            while os.path.exists(dest):
                dest = os.path.join(destino_base, f"{base} ({count})")
                count += 1
            shutil.copytree(ruta_encontrada, dest)
            return ("COPIED", nombre_original)
        except Exception as e:
            return ("ERROR", nombre_original, str(e))

    def trabajo_en_hilo(self, nombres_carpetas, ruta_base, destino):
        # Fase 1: búsqueda
        buscados = {n.lower() for n in nombres_carpetas}
        encontrados = {}
        self.queue.put(("progress_mode", "indeterminate"))
        self.queue.put(("progress_start", 10))
        self.queue.put(("status", f"Buscando {len(buscados)} carpetas..."))
        try:
            for root, dirs, files in os.walk(ruta_base):
                if self.cancelar_busqueda: break
                for d in list(dirs):
                    if d.lower() in buscados:
                        encontrados.setdefault(d.lower(), []).append(os.path.join(root, d))
                        dirs.remove(d)
        except Exception as e:
            self.queue.put(("error", f"Error en escaneo: {e}"))
            return
        self.queue.put(("progress_stop", None))
        self.queue.put(("progress_mode", "determinate"))
        self.queue.put(("status", "Iniciando copia paralela..."))
        self.queue.put(("progress_config", len(nombres_carpetas)))

        # Fase 2: copia
        processed = 0
        copied = 0
        not_found = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as exec:
            futures = {}
            for n in nombres_carpetas:
                key = n.lower()
                if key in encontrados:
                    for path in encontrados[key]:
                        futures[exec.submit(self._copiar_una_carpeta, n, path, destino)] = n
                else:
                    not_found.append(n)
            # log no encontrados
            for n in not_found:
                if self.cancelar_busqueda: break
                self.queue.put(("log", f"NO ENCONTRADA: '{n}'"))
                processed += 1
                pct = int(processed / len(nombres_carpetas) * 100)
                self.queue.put(("progress_value", processed))
                self.queue.put(("status", f"Procesando {processed}/{len(nombres_carpetas)}... {pct}%"))
            # procesar copias
            for fut in concurrent.futures.as_completed(futures):
                if self.cancelar_busqueda:
                    fut.cancel()
                    continue
                status, name, *err = fut.result()
                if status == "COPIED":
                    self.queue.put(("log", f"COPIA FINALIZADA: '{name}'"))
                    copied += 1
                else:
                    self.queue.put(("log", f"ERROR al copiar '{name}': {err[0]}"))
                processed += 1
                pct = int(processed / len(nombres_carpetas) * 100)
                self.queue.put(("progress_value", processed))
                self.queue.put(("status", f"Procesando {processed}/{len(nombres_carpetas)}... {pct}%"))

        if self.cancelar_busqueda:
            self.queue.put(("cancelled", None))
            return

        # Fase 3: reporte
        if not_found:
            try:
                rpt = os.path.join(destino, "reporte_carpetas_no_encontradas.txt")
                with open(rpt, "w", encoding="utf-8") as f:
                    f.write("Carpetas no encontradas:\n")
                    for x in not_found:
                        f.write(f"- {x}\n")
                self.queue.put(("finished", f"Completado. {copied} copiadas. Reporte en {rpt}"))
            except Exception as e:
                self.queue.put(("finished", f"Completado, pero error al crear reporte: {e}"))
        else:
            self.queue.put(("finished", f"Completado. Todas las {copied} carpetas copiadas."))

    def check_queue(self):
        try:
            while True:
                t, d = self.queue.get_nowait()
                self.procesar_mensaje(t, d)
        except queue.Empty:
            pass
        self.master.after(100, self.check_queue)

    def procesar_mensaje(self, msg_type, data):
        if msg_type == "log":
            self.txt_recorrido.text.config(state=tk.NORMAL)
            self.txt_recorrido.text.insert(tk.END, data + "\n")
            self.txt_recorrido.text.see(tk.END)
            self.txt_recorrido.text.config(state=tk.DISABLED)
        elif msg_type == "status":
            self.progress_text.set(data)
        elif msg_type == "progress_mode":
            self.barra_progreso.config(mode=data)
        elif msg_type == "progress_start":
            self.barra_progreso.start(data or 10)
        elif msg_type == "progress_stop":
            self.barra_progreso.stop()
        elif msg_type == "progress_config":
            self.barra_progreso.config(value=0, maximum=data)
        elif msg_type == "progress_value":
            self.barra_progreso.config(value=data)
        elif msg_type in ("cancelled", "finished", "error"):
            self.btn_iniciar.config(state=NORMAL)
            self.btn_cancelar.config(state=DISABLED)
            self.barra_progreso.stop()
            if msg_type == "cancelled":
                messagebox.showwarning("Cancelado", "Búsqueda cancelada.")
                self.progress_text.set("Cancelado")
                self.barra_progreso.configure(bootstyle="warning")
            elif msg_type == "finished":
                messagebox.showinfo("Finalizado", data)
                self.progress_text.set("Finalizado")
                if self.barra_progreso["mode"] == "determinate":
                    self.barra_progreso["value"] = self.barra_progreso["maximum"]
                self.barra_progreso.configure(bootstyle="success")
            else:
                messagebox.showerror("Error", data)
                self.progress_text.set("Error")
                self.barra_progreso.configure(bootstyle="danger")

    def leer_nombres_carpetas(self, archivo, nombre_columna_especificado=None):
        nombres, deptos = [], []
        ext = os.path.splitext(archivo)[1].lower()
        try:
            if ext in ['.xlsx', '.xls']:
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active
                headers = [str(c.value or '').strip().lower() for c in sheet[1]]
                if not any(headers):
                    messagebox.showerror("Error Planilla", f"'{os.path.basename(archivo)}' sin encabezados.")
                    return [], []
                name_idx = 0
                if nombre_columna_especificado:
                    try:
                        name_idx = headers.index(nombre_columna_especificado.lower())
                    except ValueError:
                        messagebox.showerror("Error Planilla", f"Columna '{nombre_columna_especificado}' no existe.")
                        return [], []
                dept_idx = headers.index('depto') if 'depto' in headers else None
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > name_idx and row[name_idx]:
                        nombres.append(str(row[name_idx]).strip())
                        deptos.append(str(row[dept_idx]).strip() if dept_idx is not None and len(row) > dept_idx and row[dept_idx] else '')
            elif ext == '.csv':
                with open(archivo, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    headers = [h.strip().lower() for h in next(reader)]
                    if not any(headers):
                        messagebox.showerror("Error Planilla", f"'{os.path.basename(archivo)}' sin encabezados.")
                        return [], []
                    name_idx = 0
                    if nombre_columna_especificado:
                        try:
                            name_idx = headers.index(nombre_columna_especificado.lower())
                        except ValueError:
                            messagebox.showerror("Error Planilla", f"Columna '{nombre_columna_especificado}' no existe.")
                            return [], []
                    dept_idx = headers.index('depto') if 'depto' in headers else None
                    for row in reader:
                        if row and len(row) > name_idx and row[name_idx]:
                            nombres.append(row[name_idx].strip())
                            deptos.append(row[dept_idx].strip() if dept_idx is not None and len(row) > dept_idx and row[dept_idx] else '')
            else:
                messagebox.showerror("Error Archivo", f"Formato no soportado: {ext}")
                return [], []
        except Exception as e:
            messagebox.showerror("Error Lectura", str(e))
            return [], []
        return nombres, deptos


if __name__ == '__main__':
    root = ttk.Window()
    setup_theme(root)
    app = BuscadorCarpetasApp(root)
    root.mainloop()
